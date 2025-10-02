from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash, send_file, make_response
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash
from models import (
    HousingUnit, Room, Bed, Employee, Company, BedAssignment, User,
    Asset, AssetLink, AssetAction, AssetDisposal, AttendanceRecord,
    FingerprintArchive, MaintenanceRequest, MaintenanceTeam,
    Assetwarehouse, Consumable, CleaningMaterial, ExpenseItem,
    MonthlyConsumption, DailyResident, MonthlyResidentAverage,
    BedTransfer, work_type_schedule
)
from sqlalchemy import literal, literal_column, or_, func, text, extract
from sqlalchemy.orm import joinedload
from datetime import datetime, date, timedelta
from collections import defaultdict, Counter
import io
from fpdf import FPDF
import arabic_reshaper
from bidi.algorithm import get_display
import pdfkit
from calendar import month_name

bp = Blueprint('main', __name__)
from flask import render_template, redirect, url_for, flash, request, session
from flask_login import login_user, logout_user, login_required, current_user
from models import User
from openpyxl import load_workbook
import pandas as pd
    # أضف كود بديل هنا

@bp.route('/login', methods=['GET', 'POST'])
def login():
    print("=== 🔐 بدء عملية تسجيل الدخول ===")

    if current_user.is_authenticated:
        print("✅ المستخدم مسجل دخول بالفعل")
        return redirect(url_for('main.home'))

    print(f"📝 طريقة الطلب: {request.method}")

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        remember = bool(request.form.get('remember_me'))

        print(f"📨 بيانات النموذج المستلمة:")
        print(f"   👤 المستخدم: '{username}'")
        print(f"   🔑 كلمة المرور: '{password}'")
        print(f"   💾 تذكرني: {remember}")
        print(f"   📋 جميع بيانات النموذج: {dict(request.form)}")

        user = User.query.filter_by(username=username).first()

        if user:
            print(f"✅ تم العثور على المستخدم في قاعدة البيانات:")
            print(f"   🆔 ID: {user.id}")
            print(f"   👤 اسم المستخدم: {user.username}")
            print(f"   📧 الإيميل: {user.email}")
            print(f"   🟢 نشط: {user.is_active}")
            print(f"   👥 الدور: {user.role}")

            # التحقق من كلمة المرور
            password_correct = user.check_password(password)
            print(f"🔐 التحقق من كلمة المرور: {password_correct}")

            if password_correct and user.is_active:
                login_user(user, remember=remember)
                print("🎉 تم تسجيل الدخول بنجاح!")
                print(f"🔓 المستخدم الحالي: {current_user}")
                flash('تم تسجيل الدخول بنجاح!', 'success')
                return redirect(url_for('main.home'))
            else:
                if not password_correct:
                    print("❌ كلمة المرور غير صحيحة")
                if not user.is_active:
                    print("❌ الحساب غير مفعل")
        else:
            print("❌ لم يتم العثور على مستخدم بهذا الاسم")

        flash('اسم المستخدم أو كلمة المرور غير صحيحة', 'error')

    return render_template('login.html')


@bp.route('/logout')
@login_required
def logout():
    print("=== 🚪 بدء عملية تسجيل الخروج ===")
    print(f"👤 المستخدم الحالي: {current_user}")

    # تسجيل الخروج من Flask-Login
    logout_user()

    # تنظيف شامل للجلسة
    session.clear()

    # إنشاء response مع حذف الكوكيز
    from flask import make_response, redirect
    response = make_response(redirect(url_for('main.login')))

    # حذف كوكيز الجلسة
    response.set_cookie('session', '', expires=0, path='/')
    response.set_cookie('remember_token', '', expires=0, path='/')

    print("✅ تم تسجيل الخروج وتنظيف الكوكيز")
    flash('تم تسجيل الخروج بنجاح', 'info')
    return response

@bp.route('/force-logout')
def force_logout():
    print("=== 🔥 تسجيل خروج قسري ===")
    logout_user()
    session.clear()

    # حذف cookies يدوياً
    from flask import make_response, redirect
    response = make_response(redirect(url_for('main.login')))
    response.set_cookie('session', '', expires=0)

    print("✅ تم التسجيل القسري للخروج")
    flash('تم تسجيل الخروج القسري', 'info')
    return response

from models import User

from flask_login import login_required, current_user
import json

# نظام الصلاحيات
PERMISSIONS = {
    'view_dashboard': 'عرض لوحة التحكم',
    'view_reports': 'عرض التقارير',
    'manage_users': 'إدارة المستخدمين',
    'manage_rooms': 'إدارة الغرف',
    'manage_finance': 'إدارة الشؤون المالية',
    'manage_settings': 'إدارة الإعدادات',
    'view_audit_logs': 'عرض سجلات التدقيق',
    'export_data': 'تصدير البيانات',
    'manage_permissions': 'إدارة الصلاحيات'
}

ROLES = {
    'viewer': 'مشاهد',
    'user': 'مستخدم عادي',
    'moderator': 'مشرف',
    'admin': 'مسؤول'
}


@bp.route('/users')
@login_required
def users_list():
    if not current_user.can_manage_users():
        flash('غير مصرح لك بالوصول إلى هذه الصفحة', 'error')
        return redirect(url_for('main.home'))

    users = User.query.all()
    return render_template('users/list.html', users=users, ROLES=ROLES)


@bp.route('/users/add', methods=['GET', 'POST'])
@login_required
def add_user():
    if not current_user.can_manage_users():
        flash('غير مصرح لك بإضافة مستخدمين', 'error')
        return redirect(url_for('main.users_list'))

    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        full_name = request.form.get('full_name')
        phone = request.form.get('phone')
        department = request.form.get('department')
        role = request.form.get('role', 'user')

        # التحقق من عدم تكرار اسم المستخدم أو الإيميل
        if User.query.filter_by(username=username).first():
            flash('اسم المستخدم موجود مسبقاً', 'error')
            return render_template('users/add.html', ROLES=ROLES, PERMISSIONS=PERMISSIONS)

        if User.query.filter_by(email=email).first():
            flash('البريد الإلكتروني موجود مسبقاً', 'error')
            return render_template('users/add.html', ROLES=ROLES, PERMISSIONS=PERMISSIONS)

        # إنشاء المستخدم الجديد
        new_user = User(
            username=username,
            email=email,
            full_name=full_name,
            phone=phone,
            department=department,
            role=role,
            active=True
        )
        new_user.set_password(password)

        # تعيين الصلاحيات الافتراضية حسب الدور
        default_permissions = new_user.get_available_permissions()
        new_user.set_permissions(default_permissions)

        db.session.add(new_user)
        db.session.commit()

        flash('تم إضافة المستخدم بنجاح', 'success')
        return redirect(url_for('main.users_list'))

    return render_template('users/add.html', ROLES=ROLES, PERMISSIONS=PERMISSIONS)


@bp.route('/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):
    if not current_user.can_manage_users():
        flash('غير مصرح لك بتعديل المستخدمين', 'error')
        return redirect(url_for('main.users_list'))

    user = User.query.get_or_404(user_id)

    if request.method == 'POST':
        user.username = request.form.get('username')
        user.email = request.form.get('email')
        user.full_name = request.form.get('full_name')
        user.phone = request.form.get('phone')
        user.department = request.form.get('department')
        user.role = request.form.get('role')
        user.active = bool(request.form.get('active'))

        # إدارة الصلاحيات
        selected_permissions = request.form.getlist('permissions')
        user.set_permissions(selected_permissions)

        # تحديث كلمة المرور إذا تم إدخالها
        new_password = request.form.get('new_password')
        if new_password:
            user.set_password(new_password)

        db.session.commit()
        flash('تم تحديث بيانات المستخدم بنجاح', 'success')
        return redirect(url_for('main.users_list'))

    user_permissions = user.get_permissions()
    return render_template('users/edit.html', user=user, ROLES=ROLES,
                           PERMISSIONS=PERMISSIONS, user_permissions=user_permissions)


@bp.route('/users/permissions/<int:user_id>')
@login_required
def user_permissions(user_id):
    if not current_user.can_manage_users():
        return jsonify({'success': False, 'message': 'غير مصرح لك'})

    user = User.query.get_or_404(user_id)
    return jsonify({
        'success': True,
        'permissions': user.get_permissions(),
        'available_permissions': user.get_available_permissions()
    })


@bp.route('/users/update_role/<int:user_id>', methods=['POST'])
@login_required
def update_user_role(user_id):
    if not current_user.can_manage_users():
        return jsonify({'success': False, 'message': 'غير مصرح لك'})

    user = User.query.get_or_404(user_id)
    new_role = request.json.get('role')

    if new_role in ROLES:
        user.role = new_role
        # تحديث الصلاحيات تلقائياً حسب الدور الجديد
        default_permissions = user.get_available_permissions()
        user.set_permissions(default_permissions)

        db.session.commit()
        return jsonify({'success': True, 'message': 'تم تحديث الدور والصلاحيات'})

    return jsonify({'success': False, 'message': 'دور غير صحيح'})





# حذف مستخدم
@bp.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    if not current_user.has_permission('manage_users') and current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'غير مصرح لك بحذف المستخدمين'})

    user = User.query.get_or_404(user_id)

    # منع حذف المستخدم الحالي
    if user.id == current_user.id:
        return jsonify({'success': False, 'message': 'لا يمكن حذف حسابك الشخصي'})

    db.session.delete(user)
    db.session.commit()

    return jsonify({'success': True, 'message': 'تم حذف المستخدم بنجاح'})


# تفعيل/تعطيل مستخدم
@bp.route('/users/toggle/<int:user_id>', methods=['POST'])
@login_required
def toggle_user(user_id):
    if not current_user.has_permission('manage_users') and current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'غير مصرح لك بهذا الإجراء'})

    user = User.query.get_or_404(user_id)
    user.active = not user.active
    db.session.commit()

    status = "مفعل" if user.active else "معطل"
    return jsonify({'success': True, 'message': f'تم {status} المستخدم', 'active': user.active})


# صفحة الملف الشخصي
@bp.route('/profile')
@login_required
def profile():
    return render_template('users/profile.html', user=current_user)


# تحديث الملف الشخصي
@bp.route('/profile/update', methods=['POST'])
@login_required
def update_profile():
    current_user.full_name = request.form.get('full_name')
    current_user.phone = request.form.get('phone')
    current_user.department = request.form.get('department')
    current_user.email = request.form.get('email')

    # تحديث كلمة المرور إذا تم إدخالها
    new_password = request.form.get('new_password')
    if new_password:
        current_user.set_password(new_password)
        flash('تم تحديث كلمة المرور بنجاح', 'success')

    db.session.commit()
    flash('تم تحديث الملف الشخصي بنجاح', 'success')
    return redirect(url_for('main.profile'))



@bp.route('/debug/user')
def debug_user():
    user_info = {
        'is_authenticated': current_user.is_authenticated,
        'user_id': current_user.get_id() if current_user.is_authenticated else None,
        'username': current_user.username if current_user.is_authenticated else None
    }
    return user_info


@bp.route('/debug')
def debug():
    users = User.query.all()
    result = "<h1>المستخدمون في قاعدة البيانات:</h1>"
    for user in users:
        result += f"""
        <div style="border:1px solid #ccc; padding:10px; margin:10px;">
            <strong>Username:</strong> {user.username}<br>
            <strong>Email:</strong> {user.email}<br>
            <strong>Active:</strong> {user.active}<br>
            <strong>Role:</strong> {user.role}<br>
            <strong>Has Password:</strong> {bool(user.password_hash)}<br>
            <strong>Password Test:</strong> 
            <form method="POST" action="/test-password/{user.id}" style="display:inline;">
                <input type="password" name="test_password" placeholder="test password">
                <button type="submit">Test</button>
            </form>
        </div>
        """
    return result

@bp.route('/test-password/<int:user_id>', methods=['POST'])
def test_password(user_id):
    user = User.query.get(user_id)
    test_pass = request.form.get('test_password')
    if user and test_pass:
        return f"""
        Password check for {user.username}: {user.check_password(test_pass)}<br>
        <a href="/debug">Back</a>
        """
    return "Error"
# صفحة حول النظام
@bp.route('/about')
@login_required
def about():
    return render_template('about.html')


# الصفحة الرئيسية

@bp.route('/')
@login_required
def home():
    return render_template('home.html')
# --- سكنات ---

def update_fingerprint_archive_for_date(current_date):
    """
    تحديث FingerprintArchive للأيام المحددة بحيث تعكس ربط الموظفين بالسكن الصحيح.
    dates: قائمة من الكائنات datetime.date
    """
    for day in dates:
        # حذف أي أرشيف موجود لذلك اليوم لتجنب التكرار
        FingerprintArchive.query.filter(FingerprintArchive.date == day).delete()
        db.session.commit()

        # جلب جميع سجلات البصمة من المصدر (أو API/جدول الموظفين المبصمين) لذلك اليوم
        fingerprint_records = fetch_fingerprint_data(day)  # افترض أن لديك هذه الدالة

        for rec in fingerprint_records:
            emp_code = int(float(rec["employee_code"])) if isinstance(rec["employee_code"], str) else rec["employee_code"]
            emp = Employee.query.filter_by(employee_code=emp_code).first()

            if emp:
                # جلب آخر تخصيص سرير للموظف لهذا اليوم
                assignment = (
                    BedAssignment.query
                    .filter(
                        BedAssignment.employee_id == emp.id,
                        BedAssignment.active == True,
                        BedAssignment.start_date <= day,
                        (BedAssignment.end_date == None) | (BedAssignment.end_date >= day)
                    )
                    .order_by(BedAssignment.start_date.desc())
                    .first()
                )

                if assignment:
                    bed = assignment.bed
                    room = bed.room
                    housing = room.housing_unit
                    company_housing_name = housing.name
                    room_number = room.room_number
                    bed_number = bed.bed_number
                else:
                    company_housing_name = "فارغ"
                    room_number = "فارغ"
                    bed_number = "فارغ"

                archive_record = FingerprintArchive(
                    employee_code=emp.employee_code,
                    name=emp.name,
                    company=emp.company.name if emp.company else "-",
                    department=emp.department or "-",
                    job_title=emp.job_title or "-",
                    work_type=emp.work_type or "-",
                    company_housing_name=company_housing_name,
                    room_number=room_number,
                    bed_number=bed_number,
                    check_in=rec.get("check_in") or "-",
                    check_out=rec.get("check_out") or "-",
                    date=day
                )
            else:
                # موظف غير موجود
                archive_record = FingerprintArchive(
                    employee_code=emp_code,
                    name="غير موجود",
                    company="-",
                    department="-",
                    job_title="-",
                    work_type="-",
                    company_housing_name="فارغ",
                    room_number="فارغ",
                    bed_number="فارغ",
                    check_in=rec.get("check_in") or "-",
                    check_out=rec.get("check_out") or "-",
                    date=day
                )

            db.session.add(archive_record)

        db.session.commit()

@bp.route('/housing')
def housing_units ():
    units = HousingUnit.query.all()
    return render_template('housing/housing_units.html', units=units)

@bp.route('/housing/add', methods=['GET', 'POST'])
def housing_add():
    if request.method == 'POST':
        name = request.form['name']
        number = request.form['number']
        total_rooms = int(request.form.get('total_rooms', 0))
        total_beds = int(request.form.get('total_beds', 0))
        new_unit = HousingUnit(name=name, number=number, total_rooms=total_rooms, total_beds=total_beds)
        db.session.add(new_unit)
        db.session.commit()
        flash('تم إضافة السكن بنجاح', 'success')
        return redirect(url_for('main.housing_units'))
    return render_template('housing/add_edit_housing_unit.html', action="إضافة")

@bp.route('/housing/edit/<int:unit_id>', methods=['GET', 'POST'])
def housing_edit(unit_id):
    unit = HousingUnit.query.get_or_404(unit_id)
    if request.method == 'POST':
        unit.name = request.form['name']
        unit.number = request.form['number']
        unit.total_rooms = int(request.form.get('total_rooms', 0))
        unit.total_beds = int(request.form.get('total_beds', 0))
        db.session.commit()
        flash('تم تعديل السكن بنجاح', 'success')
        return redirect(url_for('main.housing_units'))
    return render_template('housing/add_edit_housing_unit.html', action="تعديل", unit=unit)


@bp.route('/housing/delete/<int:unit_id>', methods=['POST'])
def housing_delete(unit_id):
    unit = HousingUnit.query.get_or_404(unit_id)
    # يمكن إضافة تحقق لمنع الحذف إذا يوجد غرف مرتبطة بالسكن
    if unit.rooms:
        flash('لا يمكن حذف السكن لأنه يحتوي على غرف.', 'danger')
        return redirect(url_for('main.housing_units'))
    db.session.delete(unit)
    db.session.commit()
    flash('تم حذف السكن بنجاح', 'success')
    return redirect(url_for('main.housing_units'))

@bp.route('/housing/<int:housing_id>/rooms')
def housing_rooms(housing_id):
    housing = HousingUnit.query.get_or_404(housing_id)
    # لا حاجة لاستدعاء .all() لأنها قائمة جاهزة
    return render_template('housing/rooms.html', housing=housing, rooms=housing.rooms)

# --- غرف ---
@bp.route('/rooms/add/<int:housing_id>', methods=['GET', 'POST'])
def room_add(housing_id):
    housing = HousingUnit.query.get_or_404(housing_id)
    if request.method == 'POST':
        room_number = request.form['room_number']
        total_beds = int(request.form['total_beds'])
        new_room = Room(
            room_number=room_number,
            total_beds=total_beds,
            housing_unit_id=housing.id  # <-- مهم جدًا
        )
        db.session.add(new_room)
        db.session.commit()

        # إضافة الأسرة تلقائيًا حسب total_beds
        for i in range(1, total_beds+1):
            new_bed = Bed(
                bed_number=str(i),
                room_id=new_room.id
            )
            db.session.add(new_bed)
        db.session.commit()

        flash('تم إضافة الغرفة والأسرة بنجاح', 'success')
        return redirect(url_for('main.housing_rooms', housing_id=housing.id))

    return render_template('housing/add_edit_room.html', action="إضافة", housing=housing)


@bp.route('/rooms/edit/<int:room_id>', methods=['GET', 'POST'])
def room_edit(room_id):
    room = Room.query.get_or_404(room_id)
    housing = room.housing_unit
    if request.method == 'POST':
        room.room_number = request.form['room_number']
        room.total_beds = int(request.form['total_beds'])
        db.session.commit()

        # ✅ إضافة الأسرة تلقائيًا إذا لم يكن هناك أي سرير
        existing_beds = room.beds.all() if hasattr(room.beds, 'all') else room.beds  # تحقق من نوع العلاقة
        if len(existing_beds) == 0:
            for i in range(1, room.total_beds + 1):
                new_bed = Bed(
                    bed_number=str(i),
                    room_id=room.id
                )
                db.session.add(new_bed)
            db.session.commit()

        flash('تم تعديل الغرفة بنجاح', 'success')
        return redirect(url_for('main.housing_rooms', housing_id=housing.id))

    return render_template('housing/add_edit_room.html', action="تعديل", room=room, housing=housing)



@bp.route('/rooms/<int:room_id>/beds')
def room_beds(room_id):

    room = Room.query.get_or_404(room_id)
    beds = room.beds.all()  # أو list(room.beds)

    return render_template(
        'housing/beds.html',
        room=room,
        beds=beds,
        show_back_to_rooms_button=True  # ← هذا السطر يضيف الزر
    )

@bp.route('/rooms/delete/<int:room_id>', methods=['POST'])
def room_delete(room_id):
    room = Room.query.get_or_404(room_id)
    housing_id = room.housing_unit_id  # حفظ معرف السكن قبل الحذف
    db.session.delete(room)
    db.session.commit()
    flash('تم حذف الغرفة بنجاح', 'success')
    return redirect(url_for('main.housing_rooms', housing_id=housing_id))

# --- أسرة ---

@bp.route('/beds/add', methods=['GET', 'POST'])
def bed_add():
    room_id = request.args.get('room_id')
    if not room_id:
        flash("لم يتم تحديد الغرفة.", "danger")
        return redirect(url_for('main.room_beds'))  # عدل حسب راوت عرض الغرف

    room = Room.query.get(room_id)
    if not room:
        flash("الغرفة غير موجودة.", "danger")
        return redirect(url_for('main.room_beds'))

    if request.method == 'POST':
        bed_number = request.form.get('bed_number')

        current_beds_count = Bed.query.filter_by(room_id=room.id).count()
        if current_beds_count >= room.total_beds:
            flash("تم الوصول للحد الأقصى لعدد الأسرة في هذه الغرفة.", "danger")
            return redirect(url_for('bed_add') + f"?room_id={room_id}")

        new_bed = Bed(room_id=room.id, bed_number=bed_number)
        db.session.add(new_bed)
        db.session.commit()

        flash("تم إضافة السرير بنجاح.", "success")
        return redirect(url_for('main.room_beds', room_id=room_id))  # عدل حسب راوت عرض الأسرة

    # GET: هنا تأكد أن room معرف
    return render_template('housing/add_edit_bed.html', room=room, action="إضافة")

@bp.route('/beds/edit/<int:bed_id>', methods=['GET', 'POST'])
def bed_edit(bed_id):
    bed = Bed.query.get_or_404(bed_id)
    room = bed.room
    if request.method == 'POST':
        bed.bed_number = request.form['bed_number']
        db.session.commit()
        flash('تم تعديل السرير بنجاح', 'success')
        return redirect(url_for('main.room_beds', room_id=room.id))
    return render_template('housing/add_edit_bed.html', action="تعديل", bed=bed, room=room)

@bp.route('/beds/delete/<int:bed_id>', methods=['POST'])
def bed_delete(bed_id):
    bed = Bed.query.get_or_404(bed_id)
    db.session.delete(bed)
    db.session.commit()
    flash('تم حذف السرير بنجاح', 'success')
    return redirect(url_for('main.room_beds', room_id=bed.room_id))

# --- موظفين ---

# --- تخصيص الأسرة ---


from models import db, Employee, Bed, Room, HousingUnit, BedAssignment, BedTransfer, Company



def get_latest_bed_info(employee):
    """إرجاع آخر تخصيص للسرير للموظف"""
    last_assignment = employee.assignments.order_by(BedAssignment.start_date.desc()).first()
    if last_assignment:
        bed = last_assignment.bed
        room = bed.room
        housing = room.housing_unit
        return {
            "company_housing_name": housing.name,
            "room_number": room.room_number,
            "bed_number": bed.bed_number
        }
    return {"company_housing_name": "-", "room_number": "-", "bed_number": "-"}
def get_employee_with_bed(emp_code):
    """إرجاع بيانات الموظف مع آخر تخصيص نشط للسرير"""
    emp = Employee.query.options(
        joinedload(Employee.assignments)
        .joinedload(BedAssignment.bed)
        .joinedload(Bed.room)
        .joinedload(Room.housing_unit)
    ).filter(Employee.employee_code == emp_code).first()

    if not emp:
        return None

    # آخر تخصيص نشط
    last_assignment = next((a for a in sorted(emp.assignments, key=lambda x: x.start_date, reverse=True) if a.active), None)

    if last_assignment:
        bed = last_assignment.bed
        room = bed.room if bed else None
        housing = room.housing_unit if room else None
        bed_info = {
            "company_housing_name": housing.name if housing else "فارغ",
            "room_number": room.room_number if room else "فارغ",
            "bed_number": bed.bed_number if bed else "فارغ",
            "assignment_type": last_assignment.assignment_type,
            "start_date": last_assignment.start_date.strftime('%Y-%m-%d') if last_assignment.start_date else "-",
            "end_date": last_assignment.end_date.strftime('%Y-%m-%d') if last_assignment.end_date else "-"
        }
    else:
        bed_info = {
            "company_housing_name": "فارغ",
            "room_number": "فارغ",
            "bed_number": "فارغ",
            "assignment_type": "-",
            "start_date": "-",
            "end_date": "-"
        }

    # استخدام الحقول الموجودة فقط في جدول Employee
    return {
        "employee_code": emp.employee_code,
        "name": emp.name,
        "job_title": emp.job_title or "-",
        "department": emp.department or "-",
        "work_type": emp.work_type or "-",  # استخدام work_type فقط
        "company_name": emp.company_name or "-",
        "employee_housing_location": emp.employee_housing_location or "-",
        "service_type": emp.service_type or "غير محدد",  # ✅ نوع الخدمة
        "company_housing_location": emp.company_housing_location or "-",
        "company_housing_name": emp.company_housing_name or "-",
        "room_number": emp.room_number or "-",
        "bed_number": emp.bed_number or "-",
        "presence_in_company": emp.presence_in_company,
        **bed_info
    }
@bp.route('/employees_report')
def employees_report():
    employees_list = [get_employee_with_bed(emp.employee_code) for emp in Employee.query.all()]

    # تصفية القيم None
    employees_list = [emp for emp in employees_list if emp is not None]

    # تجميع الموظفين حسب الشركة
    employees_by_company = defaultdict(list)
    for emp in employees_list:
        company_name = emp.get("company_name", "غير محدد")
        employees_by_company[company_name].append(emp)

    company_summary = [{"company_name": k, "count": len(v)} for k, v in employees_by_company.items()]
    total_count = sum(len(v) for v in employees_by_company.values())

    return render_template(
        "housing/employees_report.html",
        employees=employees_list,
        company_summary=company_summary,
        total_count=total_count,
        employees_by_company=employees_by_company
    )

@bp.route('/add_edit_employees_all', methods=['GET', 'POST'])
def add_edit_employees_all():
    if request.method == 'POST':
        try:
            # جمع بيانات النموذج
            data = {
                "employee_code": request.form['employee_code'],
                "name": request.form['name'],
                "job_title": request.form.get('job_title', ''),
                "department": request.form.get('department', ''),
                "work_type": request.form.get('work_type_schedule', ''),  # استخدام work_type_schedule من النموذج لملء work_type في الجدول
                "company_id": request.form.get('company_id', None),
                "company_name": request.form.get('company_name', ''),
                "employee_housing_location": request.form.get('employee_housing_location', ''),
                "service_type": request.form.get('service_type', ''),
                "company_housing_location": request.form.get('company_housing_location', ''),
                "company_housing_name": request.form.get('company_housing_name', ''),
                "room_number": request.form.get('room_number', ''),
                "bed_number": request.form.get('bed_number', ''),
                "presence_in_company": int(request.form.get('presence_in_company', 1))
            }

            print(f"📥 بيانات الواردة: {data}")

            # البحث عن الموظف الحالي
            emp = Employee.query.filter_by(employee_code=data['employee_code']).first()

            if emp:
                # تحديث بيانات الموظف
                for key, value in data.items():
                    if hasattr(emp, key):
                        setattr(emp, key, value if value != '' else None)
                message = "✅ تم تحديث بيانات الموظف بنجاح"
                print(f"🔄 تم تحديث الموظف: {emp.name}")
            else:
                # إضافة موظف جديد
                emp = Employee(**data)
                db.session.add(emp)
                message = "✅ تم إضافة الموظف الجديد بنجاح"
                print(f"🆕 تم إضافة موظف جديد: {emp.name}")

            db.session.commit()

            # تحديث تخصيص السرير بعد إضافة/تحديث الموظف
            if data['bed_number'] and data['room_number'] and data['company_housing_name']:
                print(f"🛏️ محاولة تحديث تخصيص السرير...")
                # البحث عن السرير عبر العلاقات
                bed = Bed.query.join(Room).join(HousingUnit).filter(
                    Bed.bed_number == data['bed_number'],
                    Room.room_number == data['room_number'],
                    HousingUnit.name == data['company_housing_name']
                ).first()

                if bed:
                    print(f"✅ تم العثور على السرير: {bed.id}")
                    # إلغاء أي تخصيص نشط حالي للموظف
                    current_assignment = BedAssignment.query.filter_by(employee_id=emp.id, active=True).first()
                    if current_assignment:
                        current_assignment.active = False
                        print(f"🔴 تم إلغاء التخصيص السابق: {current_assignment.id}")

                    # إنشاء تخصيص جديد للسرير
                    new_assignment = BedAssignment(
                        employee_id=emp.id,
                        bed_id=bed.id,
                        start_date=datetime.today().date(),
                        assignment_type='permanent',
                        active=True
                    )
                    db.session.add(new_assignment)
                    db.session.commit()
                    print(f"🟢 تم إنشاء تخصيص جديد: {new_assignment.id}")

                    # تحديث بيانات السكن
                    get_latest_bed_info(emp)

            return jsonify({"success": True, "message": message})

        except Exception as e:
            print(f"💥 خطأ في add_edit_employees_all: {str(e)}")
            db.session.rollback()
            return jsonify({"success": False, "error": str(e)})

    # عرض النموذج
    companies = Company.query.all()
    return render_template('housing/add_edit_employees_all.html', companies=companies)

@bp.route('/get_employee/<employee_code>')
def get_employee(employee_code):
    try:
        print(f"🔍 البحث عن الموظف بالرقم: {employee_code}")

        # تنظيف الرقم الوظيفي
        employee_code = employee_code.strip()

        emp = Employee.query.filter_by(employee_code=employee_code).first()

        if emp:
            print(f"✅ تم العثور على الموظف: {emp.name} (ID: {emp.id})")
            print(f"📋 بيانات الموظف - work_type: {emp.work_type}")

            # إنشاء بيانات الموظف مباشرة من النموذج
            data = {
                "employee_code": emp.employee_code,
                "name": emp.name or "",
                "job_title": emp.job_title or "",
                "department": emp.department or "",
                "work_type": emp.work_type or "",  # الحقل الموجود في الجدول
                "work_type_schedule": emp.work_type or "",  # استخدام work_type بدلاً من work_type_schedule
                "company_name": emp.company_name or "",
                "employee_housing_location": emp.employee_housing_location or "",
                "service_type": emp.service_type or "",
                "company_housing_location": emp.company_housing_location or "",
                "company_housing_name": emp.company_housing_name or "",
                "room_number": emp.room_number or "",
                "bed_number": emp.bed_number or "",
                "presence_in_company": emp.presence_in_company if emp.presence_in_company is not None else 1,
            }

            print(f"📊 بيانات المرسلة: {data}")
            return jsonify({"success": True, "employee": data})
        else:
            print(f"❌ الموظف غير موجود: {employee_code}")
            # الحصول على قائمة بالموظفين الموجودين للتصحيح
            all_employees = Employee.query.all()
            employee_codes = [str(e.employee_code) for e in all_employees if e.employee_code]
            print(f"📋 الموظفون الموجودون: {employee_codes}")
            return jsonify({"success": False, "error": "الموظف غير موجود"})

    except Exception as e:
        print(f"💥 خطأ في get_employee: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})


@bp.route('/delete_employee/<employee_code>', methods=['POST'])
def delete_employee(employee_code):
    try:
        emp = Employee.query.filter_by(employee_code=employee_code).first()
        if emp:
            # حذف أي تخصيص سرير مرتبط
            BedAssignment.query.filter_by(employee_id=emp.id).delete()
            db.session.delete(emp)
            db.session.commit()
            return jsonify({"success": True})
        else:
            return jsonify({"success": False, "error": "الموظف غير موجود"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})



# --- دوال مساعدة ---
def can_assign_employee_to_bed(bed_id, start_date=None, end_date=None, assignment_type=None):
    """التحقق من إمكانية تعيين الموظف على السرير حسب نوع التعيين والفترة"""
    existing_assignments = BedAssignment.query.filter(
        BedAssignment.bed_id == bed_id,
        BedAssignment.active == True
    ).all()

    for a in existing_assignments:
        if a.assignment_type == "ثابت" or assignment_type == "ثابت":
            return False
        if assignment_type == "تناوب" and a.assignment_type == "تناوب":
            if not ((end_date and end_date <= a.start_date) or (start_date and a.end_date and start_date >= a.end_date)):
                return False
    return True

def is_employee_already_assigned(employee_id, start_date, end_date=None, exclude_assignment_id=None):
    """التحقق من عدم وجود تخصيص آخر للموظف في نفس الفترة"""
    query = BedAssignment.query.filter(
        BedAssignment.employee_id == employee_id,
        or_(
            BedAssignment.end_date == None,
            BedAssignment.end_date >= start_date
        ),
        BedAssignment.start_date <= (end_date or date.max)
    )
    if exclude_assignment_id:
        query = query.filter(BedAssignment.id != exclude_assignment_id)
    return query.first() is not None


# --- عرض التعيينات حسب الشركة ---
@bp.route("/assignments")
def assignments():
    companies = Company.query.order_by(Company.name).all()
    selected_company_id = request.args.get("company_id", type=int)
    employees_data = []

    selected_company = None
    if selected_company_id:
        selected_company = Company.query.get(selected_company_id)

    if selected_company:
        # جلب الموظفين الذين ينتمون للشركة المحددة بالاسم
        employees = Employee.query.options(
            joinedload(Employee.assignments)
            .joinedload(BedAssignment.bed)
            .joinedload(Bed.room)
            .joinedload(Room.housing_unit)
        ).filter(Employee.company_name == selected_company.name).all()

        for e in employees:
            # الحصول على آخر تخصيص
            last_assignment = max(e.assignments, key=lambda a: a.start_date, default=None)

            employees_data.append({
                "employee_id": e.id,
                "assignment_id": last_assignment.id if last_assignment else None,
                "employee_code": e.employee_code,
                "employee_name": e.name,
                "job_title": e.job_title,
                "department": e.department,
                "work_type": e.work_type,
                "employee_company": e.company_name,  # استخدام اسم الشركة مباشرة
                "housing_name": last_assignment.bed.room.housing_unit.name if last_assignment else "-",
                "room_number": last_assignment.bed.room.room_number if last_assignment else "-",
                "bed_number": last_assignment.bed.bed_number if last_assignment else "-",
                "start_date": last_assignment.start_date if last_assignment else "-",
                "end_date": last_assignment.end_date if last_assignment else "-",
                "assignment_type": last_assignment.assignment_type if last_assignment else "-",
                "status": last_assignment.status if last_assignment else Employee.get_employee_status(
                    last_assignment.start_date if last_assignment else None,
                    e.work_type
                ),
                "bed_id": last_assignment.bed_id if last_assignment else None
            })

    all_beds = Bed.query.options(
        joinedload(Bed.room)
        .joinedload(Room.housing_unit)
    ).all()

    return render_template(
        "housing/assignments.html",
        companies=companies,
        selected_company_id=selected_company_id,
        employees=employees_data,
        all_beds=all_beds
    )

@bp.route("/assignments_by_company/<int:company_id>")
def assignments_by_company(company_id):
    employees_data = []

    employees = Employee.query.filter_by(company_id=company_id).order_by(Employee.employee_code).all()

    for e in employees:
        last_assignment = e.assignments.order_by(BedAssignment.start_date.desc()).first()

        housing_name = last_assignment.bed.room.housing_unit.name if last_assignment else "-"
        room_number = last_assignment.bed.room.room_number if last_assignment else "-"
        bed_number = last_assignment.bed.bed_number if last_assignment else "-"

        employees_data.append({
            "employee_id": e.id,
            "employee_code": e.employee_code,
            "employee_name": e.name,
            "job_title": e.job_title,
            "department": e.department,
            "work_type": e.work_type,
            "employee_company": e.company_name or "-",  # التعديل هنا
            "housing_name": housing_name,
            "room_number": room_number,
            "bed_number": bed_number,
            "start_date": last_assignment.start_date if last_assignment else "-",
            "end_date": last_assignment.end_date if last_assignment else "-",
            "assignment_type": last_assignment.assignment_type if last_assignment else "-",
            "status": last_assignment.status if last_assignment else Employee.get_employee_status(
                last_assignment.start_date if last_assignment else None,
                e.work_type
            ),
            "bed_id": last_assignment.bed_id if last_assignment else None
        })

    return render_template(
        "housing/employees_table.html",
        employees=employees_data,
        all_beds=Bed.query.all()
    )
# --- إضافة تخصيص سرير ---
@bp.route('/beds/assign/add', methods=['GET', 'POST'])
def bed_assign_add():
    if request.method == 'POST':
        employee_id = int(request.form.get('employee_id'))
        bed_id = int(request.form.get('bed_id'))
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date') or None
        assignment_type = request.form.get('assignment_type')

        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
        except ValueError:
            flash('تاريخ غير صالح.', 'danger')
            return redirect(url_for('main.bed_assign_add'))

        if not can_assign_employee_to_bed(bed_id, start_date_obj, end_date_obj, assignment_type):
            flash('لا يمكن إضافة تخصيص. السرير محجوز حسب شروط التخصيص.', 'danger')
            return redirect(url_for('main.bed_assign_add'))

        if is_employee_already_assigned(employee_id, start_date_obj, end_date_obj):
            flash('الموظف مرتبط بالفعل بسرير في نفس الفترة.', 'danger')
            return redirect(url_for('main.bed_assign_add'))

        new_assignment = BedAssignment(
            employee_id=employee_id,
            bed_id=bed_id,
            start_date=start_date_obj,
            end_date=end_date_obj,
            assignment_type=assignment_type,
            active=True
        )
        db.session.add(new_assignment)
        db.session.commit()

        # تحديث FingerprintArchive لكل يوم ضمن الفترة
        current_date = start_date_obj
        last_date = end_date_obj or start_date_obj
        while current_date <= last_date:
            update_fingerprint_archive_for_date(current_date)
            current_date += timedelta(days=1)

        flash("✅ تم إضافة التخصيص بنجاح", "success")
        return redirect(url_for('main.assignments'))

    beds = Bed.query.options(
        joinedload(Bed.room).joinedload(Room.housing_unit)
    ).all()
    employees = Employee.query.all()
    for bed in beds:
        setattr(bed, 'is_available', can_assign_employee_to_bed(bed.id))
    return render_template('housing/assign_bed_add.html', beds=beds, employees=employees)

# --- تعديل تخصيص ---
# --- تعديل تخصيص ---
@bp.route('/beds/assign/edit/<int:assignment_id>', methods=['GET', 'POST'])
def bed_assign_edit(assignment_id):
    assignment = BedAssignment.query.get_or_404(assignment_id)
    if request.method == 'POST':
        employee_id = int(request.form.get('employee_id'))
        bed_id = int(request.form.get('bed_id'))
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date') or None
        assignment_type = request.form.get('assignment_type')

        start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None

        if not can_assign_employee_to_bed(bed_id, start_date_obj, end_date_obj, assignment_type):
            flash('لا يمكن تعديل التخصيص. السرير محجوز.', 'danger')
            return redirect(url_for('main.bed_assign_edit', assignment_id=assignment_id))

        if is_employee_already_assigned(employee_id, start_date_obj, end_date_obj, exclude_assignment_id=assignment.id):
            flash('الموظف مرتبط بالفعل بسرير آخر في نفس الفترة.', 'danger')
            return redirect(url_for('main.bed_assign_edit', assignment_id=assignment_id))

        assignment.employee_id = employee_id
        assignment.bed_id = bed_id
        assignment.start_date = start_date_obj
        assignment.end_date = end_date_obj
        assignment.assignment_type = assignment_type
        db.session.commit()

        # تحديث FingerprintArchive لكل يوم ضمن الفترة
        current_date = start_date_obj
        last_date = end_date_obj or start_date_obj
        while current_date <= last_date:
            update_fingerprint_archive_for_date(current_date)
            current_date += timedelta(days=1)

        flash('✅ تم تعديل التخصيص بنجاح', 'success')
        return redirect(url_for('main.assignments'))

    beds = Bed.query.options(
        joinedload(Bed.room).joinedload(Room.housing_unit)
    ).all()
    employees = Employee.query.all()
    return render_template('housing/bed_assign_edit.html', assignment=assignment, beds=beds, employees=employees)


# --- نقل موظف ---
@bp.route('/bed/transfer/<int:assignment_id>', methods=['POST'])
def transfer_employee(assignment_id):
    assignment = BedAssignment.query.get_or_404(assignment_id)
    new_bed_id = int(request.form.get('new_bed_id'))
    if assignment.bed_id == new_bed_id:
        flash('لا يمكن النقل إلى نفس السرير.', 'warning')
        return redirect(url_for('main.assignments'))

    # إنهاء التخصيص الحالي
    old_start_date = assignment.start_date
    old_end_date = assignment.end_date or old_start_date
    assignment.end_date = datetime.utcnow()
    db.session.add(assignment)

    # إنشاء تخصيص جديد للسرير الجديد
    new_assignment = BedAssignment(
        employee_id=assignment.employee_id,
        bed_id=new_bed_id,
        start_date=datetime.utcnow(),
        assignment_type=assignment.assignment_type,
        active=True
    )
    db.session.add(new_assignment)

    # تسجيل النقل
    transfer_record = BedTransfer(
        employee_id=assignment.employee_id,
        from_bed_id=assignment.bed_id,
        to_bed_id=new_bed_id,
        transfer_date=datetime.utcnow(),
        transfer_type='نقل'
    )
    db.session.add(transfer_record)
    db.session.commit()

    # تحديث FingerprintArchive لكل يوم ضمن الفترة القديمة والجديدة
    current_date = old_start_date
    while current_date <= old_end_date:
        update_fingerprint_archive_for_date(current_date)
        current_date += timedelta(days=1)
    update_fingerprint_archive_for_date(datetime.utcnow().date())

    flash('✅ تم النقل بنجاح', 'success')
    return redirect(url_for('main.assignments'))


# --- إلغاء التخصيص ---
@bp.route('/bed_assignments/<int:assignment_id>/cancel', methods=['POST'])
def cancel_bed_assignment(assignment_id):
    assignment = BedAssignment.query.get_or_404(assignment_id)
    start_date_obj = assignment.start_date
    end_date_obj = assignment.end_date or start_date_obj
    assignment.active = False
    db.session.commit()

    # تحديث FingerprintArchive لكل يوم ضمن الفترة الملغاة
    current_date = start_date_obj
    while current_date <= end_date_obj:
        update_fingerprint_archive_for_date(current_date)
        current_date += timedelta(days=1)

    flash('✅ تم إلغاء التخصيص بنجاح.', 'success')
    return redirect(url_for('main.assignments'))


@bp.route('/bed_assignments/<int:assignment_id>/delete', methods=['POST'])
def delete_bed_assignment(assignment_id):
    assignment = BedAssignment.query.get_or_404(assignment_id)
    db.session.delete(assignment)
    db.session.commit()
    update_fingerprint_archive_for_date(date)

    flash('✅ تم حذف التخصيص.', 'success')
    return redirect(url_for('main.assignments'))


# --- تقرير النقل ---

@bp.route("/transfers_report")
def transfers_report():
    transfers = BedTransfer.query.options(
        joinedload(BedTransfer.employee),
        joinedload(BedTransfer.from_bed).joinedload(Bed.room).joinedload(Room.housing_unit),
        joinedload(BedTransfer.to_bed).joinedload(Bed.room).joinedload(Room.housing_unit)
    ).all()
    return render_template("housing/transfers_report.html", transfers=transfers)

# --- بحث عن الموظف ---

@bp.route('/search_employee')
def search_employee():
    q = request.args.get('q', '').strip()
    results = []
    if q:
        employees = Employee.query.filter(
            (Employee.name.ilike(f"%{q}%")) | (Employee.employee_code.like(f"{q}%"))
        ).all()
        for emp in employees:
            if emp.name:
                results.append({'id': emp.id, 'code': emp.employee_code, 'name': emp.name})
    return jsonify(results)


@bp.route('/reports/housing')
def housing_report():
    housings = HousingUnit.query.all()

    for housing in housings:
        housing.rooms_count_calc = len(housing.rooms)
        housing.total_beds_calc = sum(room.total_beds for room in housing.rooms)
        housing.occupied_beds_calc = sum(room.occupied_beds for room in housing.rooms)
        housing.available_beds_calc = housing.total_beds_calc - housing.occupied_beds_calc




    total_rooms = sum(housing.rooms_count_calc for housing in housings)
    total_beds = sum(housing.total_beds_calc for housing in housings)
    total_occupied = sum(housing.occupied_beds_calc for housing in housings)
    total_vacant = sum(housing.available_beds_calc for housing in housings)

    return render_template(
        'housing/housing_report.html',
        housings=housings,
        total_rooms=total_rooms,
        total_beds=total_beds,
        total_occupied=total_occupied,
        total_vacant=total_vacant
    )



@bp.route("/housing_report")
def housing_report_view():
    housings = report_hou_emp()
    return render_template("housing/report_hou_emp.html", housings=housings)

@bp.route("/housing_report/export/excel")
def housing_report_excel():
    housings = report_hou_emp()  # استدعاء الدالة التي ترجع بيانات السكنات والموظفين
    return export_hou_emp_excel(housings)



@bp.route("/housing_report/export/pdf")
def housing_report_pdf():
    housings = report_hou_emp()  # استدعاء الدالة التي ترجع بيانات السكنات والموظفين
    return export_hou_emp_pdf(housings)  # تمرير البيانات إلى دالة التصدير

import pdfkit

@bp.route('/beds/report', methods=['GET'])
def beds_report():
    # --- قراءة الفلاتر ---
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    housing_filter = request.args.get('housing_unit')

    start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date() if start_date_str else None
    end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date() if end_date_str else None
    today = datetime.utcnow().date()

    # --- إعادة تحميل الجلسة للتأكد من عدم وجود بيانات قديمة ---
    db.session.expire_all()

    # --- جلب السكنات والأسرة ---
    housing_units = HousingUnit.query.order_by(HousingUnit.name).all()
    beds = Bed.query.join(Bed.room).join(Room.housing_unit)\
        .order_by(HousingUnit.name, Room.room_number, Bed.bed_number).all()

    report_data = []
    overall_summary = {'total_beds': 0, 'total_employees': 0}

    # --- الحصول على آخر تخصيص لكل موظف نشط ---
    subq = db.session.query(
        BedAssignment.employee_id,
        func.max(BedAssignment.start_date).label('last_start')
    ).group_by(BedAssignment.employee_id).subquery()

    last_assignments = db.session.query(BedAssignment)\
        .join(subq, (BedAssignment.employee_id == subq.c.employee_id) &
                     (BedAssignment.start_date == subq.c.last_start))\
        .filter(BedAssignment.active == True)\
        .filter(or_(BedAssignment.end_date == None, BedAssignment.end_date >= today))\
        .all()

    # dict: bed_id -> list of last assignments فيه فقط آخر تخصيص لكل موظف
    bed_to_last_assignments = {}
    for a in last_assignments:
        bed_to_last_assignments.setdefault(a.bed_id, []).append(a)

    for housing in housing_units:
        if housing_filter and housing.name != housing_filter:
            continue

        housing_beds = [bed for bed in beds if bed.room.housing_unit_id == housing.id]
        beds_list = []
        housing_summary = {'total_beds': 0, 'total_employees': 0}
        ai_suggestions = []

        for bed in housing_beds:
            emp_list = []
            last_assigns_for_bed = bed_to_last_assignments.get(bed.id, [])

            for a in last_assigns_for_bed:
                employee = a.employee
                last_transfer = BedTransfer.query.filter_by(employee_id=a.employee_id) \
                    .order_by(BedTransfer.transfer_date.desc()).first()

                emp_list.append({
                    'employee_code': employee.employee_code if employee else "-",
                    'employee_name': employee.name if employee else "-",
                    'start_date': a.start_date,
                    'end_date': a.end_date,
                    'assignment_type': a.assignment_type,
                    'work_type': employee.work_type if employee else 'غير محدد',
                    'rotation_info': '',
                    'last_transfer': last_transfer
                })

            if len(emp_list) > 1:
                ai_suggestions.append(
                    f"السرير {bed.bed_number} بالغرفة {bed.room.room_number} يحتوي على أكثر من موظف. تحقق من جدول التناوب."
                )

            beds_list.append({'bed': bed, 'assignments': emp_list})
            housing_summary['total_beds'] += 1
            housing_summary['total_employees'] += len(emp_list)

        report_data.append({
            'housing': housing,
            'beds': beds_list,
            'summary': housing_summary,
            'ai_suggestions': ai_suggestions
        })

        overall_summary['total_beds'] += housing_summary['total_beds']
        overall_summary['total_employees'] += housing_summary['total_employees']

    return render_template(
        'housing/beds_report.html',
        report_data=report_data,
        overall_summary=overall_summary,
        start_date=start_date_str,
        end_date=end_date_str,
        housing_filter=housing_filter,
        housing_units=housing_units
    )

@bp.route('/beds/report/pdf', methods=['GET'])
def beds_report_pdf():
    # استدعاء نفس بيانات التقرير
    report_data, overall_summary = generate_report_data()  # ضع هنا دالة توليد البيانات
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    housing_filter = request.args.get('housing_unit')

    # توليد HTML من القالب
    rendered_html = render_template(
        'housing/assign_bed_report.html',
        report_data=report_data,
        overall_summary=overall_summary,
        start_date=start_date,
        end_date=end_date,
        housing_filter=housing_filter
    )

    # توليد PDF باستخدام pdfkit
    pdf_file = pdfkit.from_string(
        rendered_html,
        False,  # False لإرجاع Bytes بدلاً من ملف
        options={
            'page-size': 'A4',
            'encoding': "UTF-8",
            'enable-local-file-access': None
        }
    )

    # إرجاع الملف كـ download
    return send_file(
        io.BytesIO(pdf_file),
        as_attachment=True,
        download_name="beds_report.pdf",
        mimetype='application/pdf'
    )
@bp.route('/beds/report/excel', methods=['GET'])
def beds_report_excel():
    ...


def report_hou_emp():
    today = datetime.today().date()
    housings_data = []

    housings = HousingUnit.query.order_by(HousingUnit.name).all()

    for housing in housings:
        housing_info = {
            "name": housing.name,
            "rooms": []
        }

        for room in housing.rooms:
            for bed in room.beds.all():
                assignment = bed.assignments.filter(
                    BedAssignment.active == True,
                    BedAssignment.start_date <= today,
                    (BedAssignment.end_date == None) | (BedAssignment.end_date >= today)
                ).first()

                if assignment and assignment.employee:
                    emp_status = Employee.get_employee_status(
                        assignment.start_date,
                        assignment.employee.work_type
                    )
                    if emp_status == "نشط":
                        bed_status = "مشغول"
                        employee_name = assignment.employee.name
                        job_title = assignment.employee.job_title or ""
                        department = assignment.employee.department or ""
                    else:
                        bed_status = "فارغ"
                        employee_name = ""
                        job_title = ""
                        department = ""
                else:
                    bed_status = "فارغ"
                    employee_name = ""
                    job_title = ""
                    department = ""

                housing_info["rooms"].append({
                    "room_number": room.room_number,
                    "bed_number": bed.bed_number,
                    "status": bed_status,
                    "employee_name": employee_name,
                    "job_title": job_title,
                    "department": department
                })

        housings_data.append(housing_info)

    return housings_data


def export_hou_emp_pdf(housings):
    pdf = PDF()

    # إضافة الخطوط مرة واحدة
    pdf.add_font("Amiri", "", "D:/ghith/NEW/rooms/fonts/Amiri-Regular.ttf", uni=True)
    pdf.add_font("Amiri", "B", "D:/ghith/NEW/rooms/fonts/Amiri-Bold.ttf", uni=True)
    pdf.add_font("Amiri", "I", "D:/ghith/NEW/rooms/fonts/Amiri-Italic.ttf", uni=True)

    pdf.add_page()

    headers = ["الموظف الحالي", "المسمى الوظيفي", "القسم", "الغرفة", "رقم السرير", "الحالة"]
    col_widths = [55, 45, 30, 20, 20, 20]

    for h in housings:
        # عنوان السكن
        pdf.set_font("Amiri", "B", 14)
        pdf.cell(0, 10, ar(f"السكن: {h.get('name', 'غير محدد')}"), ln=True, align='C')
        pdf.ln(2)

        # طباعة ترويسة الجدول
        pdf.table_header(headers, col_widths)

        pdf.set_font("Amiri", "", 12)
        first_row = True
        for bed in h.get("rooms", []):
            if pdf.get_y() > 260:
                pdf.add_page()
                pdf.table_header(headers, col_widths)

            fill_row = False  # حذف تلوين الصف الأول

            pdf.cell(col_widths[0], 10, ar(bed["employee_name"]), border=1, align='C', fill=fill_row)
            pdf.cell(col_widths[1], 10, ar(bed["job_title"]), border=1, align='C', fill=fill_row)
            pdf.cell(col_widths[2], 10, ar(bed["department"]), border=1, align='C', fill=fill_row)
            pdf.cell(col_widths[3], 10, ar(bed["room_number"]), border=1, align='C', fill=fill_row)
            pdf.cell(col_widths[4], 10, str(bed["bed_number"]), border=1, align='C', fill=fill_row)

            # تلوين الحالة فقط
            if bed["status"] == "مشغول":
                pdf.set_text_color(255, 255, 255)
                pdf.set_fill_color(220, 53, 69)
                pdf.cell(col_widths[5], 10, ar(bed["status"]), border=1, align='C', fill=True)
            else:
                pdf.set_text_color(0, 0, 0)
                pdf.set_fill_color(144, 238, 144)
                pdf.cell(col_widths[5], 10, ar(bed["status"]), border=1, align='C', fill=True)

            pdf.set_text_color(0, 0, 0)
            pdf.ln()


    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="housing_report.pdf", mimetype="application/pdf")


# تصدير Excel
def export_hou_emp_excel(housings):
    rows = []
    for h in housings:
        for r in h["rooms"]:
            for b in r.get("beds", []):
                emp = b.current_employee()
                rows.append({
                    "السكن": h["name"],
                    "اسم الموظف": emp.name if emp else "لا يوجد",
                    "كود الموظف": emp.employee_code if emp else "",
                    "المسمى الوظيفي": emp.job_title if emp else "",
                    "القسم": emp.department if emp else "",
                    "الغرفة": r.get("room_number"),
                    "رقم السرير": b.bed_number,
                    "الحالة": "مشغول" if b.is_occupied else "فارغ"
                })
    df = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name="السكنات")
    output.seek(0)
    return send_file(output, as_attachment=True, download_name="housing_report.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



from fpdf import FPDF


def ar(text):
    import arabic_reshaper
    from bidi.algorithm import get_display
    if not text:
        return ""
    reshaped_text = arabic_reshaper.reshape(str(text))
    return get_display(reshaped_text)

class PDF(FPDF):
    def header(self):
        # شعار الشركة
        # self.image("static/images/logo.png", 10, 8, 33)  # إذا أردت الشعار
        self.set_font("Amiri", "B", 16)
        self.cell(0, 10, ar("التقرير المتقدم للسكنات"), border=0, ln=1, align="C")
        # التاريخ
        self.set_font("Amiri", "", 12)
        self.cell(0, 10, ar(f"تاريخ الإنشاء: {datetime.now().strftime('%Y-%m-%d')}"), border=0, ln=1, align="C")
        self.ln(5)

def housing_report_advanced_pdf(housings, analysis):
    import io
    from fpdf import FPDF
    import arabic_reshaper
    from bidi.algorithm import get_display

    def ar(text):
        if not text:
            return ""
        return get_display(arabic_reshaper.reshape(str(text)))


    pdf = FPDF()
    pdf.add_page()
    # إضافة الخطوط
    pdf.add_font("Amiri", "", "D:/ghith/NEW/rooms/fonts/Amiri-Regular.ttf", uni=True)
    pdf.add_font("Amiri", "B", "D:/ghith/NEW/rooms/fonts/Amiri-Bold.ttf", uni=True)
    pdf.add_font("Amiri", "I", "D:/ghith/NEW/rooms/fonts/Amiri-Italic.ttf", uni=True)

    headers = ["الموظف الحالي", "المسمى الوظيفي", "القسم", "الغرفة", "رقم السرير", "الحالة"]
    col_widths = [55, 45, 30, 20, 20, 20]
    line_height = 10

    for h in housings:
        pdf.set_font("Amiri", "B", 14)
        pdf.cell(0, 10, ar(f"السكن: {h.get('name', 'غير محدد')}"), ln=True, align='C')
        pdf.ln(2)

        # ترويسة الجدول
        pdf.set_font("Amiri", "B", 12)
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], line_height, ar(header), border=1, align='C', fill=True)

        pdf.ln()

        pdf.set_font("Amiri", "", 12)
        for bed in h.get("rooms", []):
            if pdf.get_y() > 260:
                pdf.add_page()
                # إعادة طباعة الترويسة بعد الصفحة الجديدة
                pdf.set_font("Amiri", "B", 12)
                for i, header in enumerate(headers):
                    pdf.cell(col_widths[i], line_height, ar(header), border=1, align='C', fill=True)
                pdf.ln()
                pdf.set_font("Amiri", "", 12)

            # تعبئة الصف
            pdf.cell(col_widths[0], line_height, ar(bed["employee_name"]), border=1, align='C')
            pdf.cell(col_widths[1], line_height, ar(bed["job_title"]), border=1, align='C')
            pdf.cell(col_widths[2], line_height, ar(bed["department"]), border=1, align='C')
            pdf.cell(col_widths[3], line_height, str(bed["room_number"]), border=1, align='C')
            pdf.cell(col_widths[4], line_height, str(bed["bed_number"]), border=1, align='C')

            # تلوين الحالة
            if bed["status"] == "مشغول":
                pdf.set_text_color(255, 255, 255)
                pdf.set_fill_color(220, 53, 69)
                pdf.cell(col_widths[5], line_height, ar(bed["status"]), border=1, align='C', fill=True)
            else:
                pdf.set_text_color(0, 0, 0)
                pdf.set_fill_color(144, 238, 144)
                pdf.cell(col_widths[5], line_height, ar(bed["status"]), border=1, align='C', fill=True)

            pdf.set_text_color(0, 0, 0)
            pdf.ln()

    # --- التحليل والمقترحات ---
    pdf.ln(5)
    pdf.set_font("Amiri", "B", 14)
    pdf.cell(0, 10, ar("تحليل ومقترحات"), ln=True)

    pdf.set_font("Amiri", "", 12)
    pdf.multi_cell(0, 8, ar(f"إجمالي الأسرة: {analysis['total_beds']}"))
    pdf.multi_cell(0, 8, ar(f"المشغولة: {analysis['occupied']}"))
    pdf.multi_cell(0, 8, ar(f"الفارغة: {analysis['vacant']}"))
    pdf.multi_cell(0, 8, ar(f"نسبة الإشغال: {analysis['occupancy_rate']}%"))
    pdf.ln(2)

    if analysis.get("suggestions"):
        pdf.set_font("Amiri", "I", 12)
        for s in analysis["suggestions"]:
            pdf.multi_cell(0, 8, ar(f"- {s}"))

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)
    return output

def build_filter_args():
    """
    تقرأ فلاتر GET:
    status: 'مشغول' | 'فارغ' | ''
    housing_id: int | ''
    room_number: str | ''
    department: str | ''
    employee_name: str | ''
    """
    return {
        "status": request.args.get("status", "").strip(),
        "housing_id": request.args.get("housing_id", "").strip(),
        "room_number": request.args.get("room_number", "").strip(),
        "department": request.args.get("department", "").strip(),
        "employee_name": request.args.get("employee_name", "").strip(),
    }


from collections import defaultdict

@bp.route("/housing_dashboard")
def housing_dashboard():
    # جلب كل الموظفين مع الشركة لتجنب lazy loading
    employees = Employee.query.all()

    # تقسيم الموظفين حسب الشركة باستخدام dict عادي
    employees_by_company = {}
    for emp in employees:
        # استخدام company_name من get_employee_with_bed لضمان ظهور الاسم الصحيح
        emp_data = get_employee_with_bed(emp.employee_code)
        company_name = emp_data.get("company_name") or "بدون شركة"

        if company_name not in employees_by_company:
            employees_by_company[company_name] = []
        employees_by_company[company_name].append(emp_data)

    # عدد الموظفين لكل شركة
    employees_count_by_company = {name: len(emps) for name, emps in employees_by_company.items()}

    # جلب بيانات السكنات
    housings = HousingUnit.query.all()
    housings_data = []

    for housing in housings:
        rooms_count = len(housing.rooms)
        total_beds = sum(room.total_beds for room in housing.rooms)
        occupied_beds = sum(room.occupied_beds for room in housing.rooms)
        empty_beds = total_beds - occupied_beds
        housings_data.append({
            "name": housing.name,
            "rooms_count": rooms_count,
            "total_beds": total_beds,
            "occupied_beds": occupied_beds,
            "empty_beds": empty_beds
        })

    # جلب جميع الأسرة
    beds = db.session.query(Bed).all()

    # احسب المشغولة بناءً على أي قيمة تمثل الاحتلال
    occupied_beds = sum(1 for b in beds if str(b.is_occupied).lower() in ['true', '1', 'yes'])

    # إجمالي الأسرة
    total_beds = len(beds)

    # الفارغة
    vacant_beds = total_beds - occupied_beds

    # نسبة الإشغال
    occupancy_rate = round((occupied_beds / total_beds) * 100, 2) if total_beds else 0

    analysis = {
        "total_beds": total_beds,
        "occupied": occupied_beds,
        "vacant": vacant_beds,
        "occupancy_rate": occupancy_rate
    }

    # جلب كل الموظفين
    employees = Employee.query.all()

    # إزالة التكرار حسب كود الموظف
    unique_employees = {}
    for emp in employees:
        unique_employees[emp.employee_code] = emp
    employees = list(unique_employees.values())
    # تجميع حسب الشركة + نوع الخدمة
    employees_by_service_company = {}
    total_by_service = {}

    for emp in employees:
        emp_data = get_employee_with_bed(emp.employee_code)
        company = emp_data.get("company_name") or "بدون شركة"
        service = emp_data.get("service_type") or "بدون خدمة"

        if company not in employees_by_service_company:
            employees_by_service_company[company] = {}
        employees_by_service_company[company][service] = employees_by_service_company[company].get(service, 0) + 1

        # إجمالي لكل خدمة عبر كل الشركات
        total_by_service[service] = total_by_service.get(service, 0) + 1

    # قائمة الشركات للاختيار في الواجهة
    companies = list(employees_by_service_company.keys())

    return render_template(
        "housing/housing_dashboard.html",
        housings=housings,
        employees_by_company=employees_count_by_company,
        housings_data=housings_data,
        analysis=analysis,
        employees_by_service=total_by_service,  # الإجمالي عبر الشركات
        employees_by_service_company=employees_by_service_company,  # لكل شركة
        companies=companies
    )

    # --- تمرير البيانات للقالب ---


@bp.route("/housing_report_html")  # غير الاسم هنا
def housing_report_html():         # وغير اسم الدالة أيضًا
    today = date.today()

    beds = Bed.query.join(Room).join(HousingUnit).all()

    report = []

    for bed in beds:
        # جلب جميع التعيينات المرتبطة بالسرير، ترتيب حسب تاريخ البدء
        assignments = bed.assignments.order_by(BedAssignment.start_date).all()

        if not assignments:
            # الأسرة فارغة
            report.append({
                "housing_name": bed.room.housing_unit.name,
                "room_number": bed.room.room_number,
                "bed_number": bed.bed_number,
                "employee_code": "",
                "employee_name": "",
                "company_name": "",
                "job_title": "",
                "department": "",
                "status": "فارغ",
                "assignment_type": ""
            })
            continue

        for assignment in assignments:
            employee = assignment.employee
            if not employee:
                continue

            # حساب الحالة بناءً على نوع الدوام وتاريخ البدء
            status = employee.get_employee_status(assignment.start_date, employee.work_type)

            # التحقق من نوع التعيين permanent أو rotating
            if assignment.assignment_type == "permanent":
                # يظهر موظف واحد فقط على السرير، الآخرين تجاهلهم
                report.append({
                    "housing_name": bed.room.housing_unit.name,
                    "room_number": bed.room.room_number,
                    "bed_number": bed.bed_number,
                    "employee_code": employee.employee_code,
                    "employee_name": employee.name,
                    "company_name": employee.company.name if employee.company else "",
                    "job_title": employee.job_title,
                    "department": employee.department,
                    "status": status,
                    "assignment_type": assignment.assignment_type
                })
                break  # لا يسمح بأي موظف آخر
            else:
                # rotating يسمح بظهور أكثر من موظف حسب نوع الدوام
                report.append({
                    "housing_name": bed.room.housing_unit.name,
                    "room_number": bed.room.room_number,
                    "bed_number": bed.bed_number,
                    "employee_code": employee.employee_code,
                    "employee_name": employee.name,
                    "company_name": employee.company.name if employee.company else "",
                    "job_title": employee.job_title,
                    "department": employee.department,
                    "status": status,
                    "assignment_type": assignment.assignment_type
                })

    # ترتيب التقرير حسب الوحدة السكنية، الغرفة، السرير
    report.sort(key=lambda x: (x['housing_name'], x['room_number'], x['bed_number']))

    return jsonify(report)


# -----------------------------
# -----------------------------
def report_hou_emp_filtered(filters=None):
    from sqlalchemy import text

    sql_str = """
    SELECT 
        e.employee_code AS employee_code,
        e.name AS employee_name,
        c.name AS company_name,
        e.job_title AS job_title,
        e.department AS department,
        h.id AS housing_id,
        h.name AS housing_name,
        r.room_number AS room_number,
        b.bed_number AS bed_number,
        CASE
            WHEN e.id IS NOT NULL THEN 'مشغول'
            ELSE 'فارغ'
        END AS status
    FROM beds b
    JOIN rooms r ON r.id = b.room_id
    JOIN housing_units h ON h.id = r.housing_unit_id
    LEFT JOIN bed_assignments ba ON ba.bed_id = b.id AND (ba.end_date IS NULL OR ba.end_date > CURRENT_DATE)
    LEFT JOIN employees e ON e.id = ba.employee_id
    LEFT JOIN companies c ON c.id = e.company_id
    WHERE 1=1
    """

    # إضافة فلاتر ديناميكية
    conditions = []
    params = {}

    if filters:
        if filters.get("status"):
            if filters["status"] == "مشغول":
                conditions.append("e.id IS NOT NULL")
            elif filters["status"] == "فارغ":
                conditions.append("e.id IS NULL")

        if filters.get("housing_id"):
            conditions.append("h.id = :housing_id")
            params["housing_id"] = filters["housing_id"]

        if filters.get("room_number"):
            conditions.append("r.room_number LIKE :room_number")
            params["room_number"] = f"%{filters['room_number']}%"

        if filters.get("department") and filters.get("status") != "فارغ":
            conditions.append("e.department LIKE :department")
            params["department"] = f"%{filters['department']}%"

        if filters.get("employee_name") and filters.get("status") != "فارغ":
            conditions.append("e.name LIKE :employee_name")
            params["employee_name"] = f"%{filters['employee_name']}%"

    if conditions:
        sql_str += " AND " + " AND ".join(conditions)

    sql_str += " ORDER BY h.name, r.room_number, b.bed_number"

    sql = text(sql_str)
    result = db.session.execute(sql, params).mappings().all()

    # تجميع البيانات بشكل هرمي: سكن → غرف → أسرة
    report = {}
    for row in result:
        housing_name = row['housing_name']
        room_number = row['room_number']

        if housing_name not in report:
            report[housing_name] = {
                "housing_name": housing_name,
                "housing_id": row['housing_id'],
                "rooms": {}
            }

        if room_number not in report[housing_name]["rooms"]:
            report[housing_name]["rooms"][room_number] = {
                "room_number": room_number,
                "beds": []
            }

        report[housing_name]["rooms"][room_number]["beds"].append({
            "bed_number": row['bed_number'],
            "employee_code": row['employee_code'] or "-",
            "employee_name": row['employee_name'] or "فارغ",
            "company_name": row['company_name'] or "-",
            "job_title": row['job_title'] or "-",
            "department": row['department'] or "-",
            "status": row['status']
        })

    # تحويل الهيكل إلى قائمة
    final_report = []
    for housing_data in report.values():
        # تحويل الغرف من قاموس إلى قائمة
        housing_data["rooms"] = list(housing_data["rooms"].values())
        final_report.append(housing_data)

    return final_report


def analyze_report(housings_data):
    """
    تحليل ذكي للقوائم المُصفّاة، وإرجاع ملخص نصي + مؤشرات.
    """
    total_beds = 0
    occupied = 0
    per_housing = {}  # {housing_name: {"total": X, "occ": Y}}
    per_department = {}  # {department: {"total": X, "occ": Y}}
    empty_rooms = []  # [(housing, room_number)] للغرف الفارغة 100%

    for housing in housings_data:
        housing_name = housing.get("housing_name", "غير محدد")
        h_total = 0
        h_occ = 0

        for room in housing.get("rooms", []):
            room_number = room.get("room_number", "غير محدد")
            room_total = 0
            room_occ = 0

            for bed in room.get("beds", []):
                total_beds += 1
                h_total += 1
                room_total += 1

                if bed.get("status") == "مشغول":
                    occupied += 1
                    h_occ += 1
                    room_occ += 1

                # تجميع حسب القسم
                dep = bed.get("department") or "غير محدد"
                per_department.setdefault(dep, {"total": 0, "occ": 0})
                per_department[dep]["total"] += 1
                if bed.get("status") == "مشغول":
                    per_department[dep]["occ"] += 1

            # التحقق من الغرف الفارغة تمامًا
            if room_total > 0 and room_occ == 0:
                empty_rooms.append((housing_name, room_number))

        # حفظ تلخيص السكن
        per_housing[housing_name] = {"total": h_total, "occ": h_occ}

    # مؤشرات عامة
    occupancy_rate = round((occupied / total_beds) * 100, 2) if total_beds else 0.0

    # اقتراحات مبنيّة على قواعد بسيطة
    suggestions = []
    if total_beds == 0:
        suggestions.append("لا توجد بيانات لعرضها. حاول تعديل عوامل التصفية.")
    else:
        if occupancy_rate < 60:
            suggestions.append(
                "انخفاض نسبة الإشغال إجمالًا؛ يُقترح إعادة توزيع الموظفين ودمج الغرف قليلة الإشغال لتقليل الهدر.")
        elif occupancy_rate > 90:
            suggestions.append(
                "نسبة إشغال مرتفعة جدًا؛ ضع خطة توسّع أسرّة أو رفع الطاقة الاستيعابية لتجنّب الاختناقات.")

        if empty_rooms:
            suggestions.append(
                f"هناك غرف فارغة بالكامل ({len(empty_rooms)} غرفة)؛ يُقترح مراجعة توزيع الأقسام أو إغلاق المؤقت لبعضها.")

        # أقسام ذات تباين
        for dep, d in per_department.items():
            if d["total"] >= 3:
                rate = (d["occ"] / d["total"]) * 100 if d["total"] else 0
                if rate < 50:
                    suggestions.append(
                        f"القسم '{dep}' إشغاله منخفض ({rate:.0f}%)؛ يمكن نقل بعض الأسرة أو إعادة التخطيط.")
                elif rate > 90:
                    suggestions.append(f"القسم '{dep}' قريب من الامتلاء ({rate:.0f}%)؛ راجع زيادة أسرة مخصّصة له.")

    analysis = {
        "total_beds": total_beds,
        "occupied": occupied,
        "vacant": total_beds - occupied,
        "occupancy_rate": occupancy_rate,
        "per_housing": per_housing,
        "per_department": per_department,
        "empty_rooms": empty_rooms,
        "suggestions": suggestions
    }
    return analysis


@bp.route("/housing_report/advanced")
def housing_report_advanced():
    all_housings = HousingUnit.query.order_by(HousingUnit.name).all()

    # بناء الفلاتر بشكل صحيح
    filters = {
        "status": request.args.get("status", ""),
        "housing_id": request.args.get("housing_id", ""),
        "room_number": request.args.get("room_number", ""),
        "department": request.args.get("department", ""),
        "employee_name": request.args.get("employee_name", "")
    }

    # تنظيف الفلاتر الفارغة
    filters = {k: v for k, v in filters.items() if v}

    data = report_hou_emp_filtered(filters)
    analysis = analyze_report(data)

    return render_template(
        "housing/report_advanced.html",
        housings=data,
        filters=filters,
        all_housings=all_housings,
        analysis=analysis,
        now=datetime.now()
    )


def build_filter_args():
    """بناء قاموس الفلاتر من request.args"""
    filters = {}

    status = request.args.get("status", "").strip()
    housing_id = request.args.get("housing_id", "").strip()
    room_number = request.args.get("room_number", "").strip()
    department = request.args.get("department", "").strip()
    employee_name = request.args.get("employee_name", "").strip()

    if status:
        filters["status"] = status
    if housing_id:
        filters["housing_id"] = housing_id
    if room_number:
        filters["room_number"] = room_number
    if department:
        filters["department"] = department
    if employee_name:
        filters["employee_name"] = employee_name

    return filters



# ---[ 2. التصدير للفلاتر نفسها ]---

@bp.route("/housing_report/advanced/export/excel")
def housing_report_advanced_excel():
    filters = build_filter_args()
    data = report_hou_emp_filtered(filters)
    return export_hou_emp_excel_for_flat(data)  # دالة مبسطة للتصدير الموضح أدناه

def export_hou_emp_excel_for_flat(housings):
    from flask import send_file

    rows = []
    for h in housings:
        for row in h["rooms"]:
            rows.append({
                "السكن": h["name"],
                "اسم الموظف": row["employee_name"],
                "المسمى الوظيفي": row["job_title"],
                "القسم": row["department"],
                "الغرفة": row["room_number"],
                "رقم السرير": row["bed_number"],
                "الحالة": row["status"],
            })
    df = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name="تقرير مُفلتر")
    output.seek(0)
    return send_file(output, as_attachment=True,
                     download_name="housing_report_filtered.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---[ تعديل تصدير PDF ليشمل التحليل والمقترحات ]---
# ---[ تعديل تصدير PDF مع التحليل والمقترحات بطريقة صحيحة ]---


@bp.route("/housing_report/advanced/export/pdf")
def housing_report_advanced_pdf_view():
    from flask import request, send_file
    # 1. جلب الفلاتر من الرابط
    filters = build_filter_args()
    # 2. استدعاء بيانات التقرير حسب الفلاتر
    data = report_hou_emp_filtered(filters)
    # 3. تحليل البيانات
    analysis = analyze_report(data)
    # 4. إنشاء PDF مع كل التنسيقات السابقة
    pdf_buffer = housing_report_advanced_pdf(data, analysis)
    # 5. إرسال الملف للمستخدم
    return send_file(
        pdf_buffer,
        as_attachment=True,
        download_name="housing_report_advanced.pdf",
        mimetype="application/pdf"
    )




# ======================== دالة لجلب بيانات البصمة ========================
@bp.route("/sync-from-local")
def sync_from_local():
    """مزامنة البيانات من البيئة المحلية إلى Railway"""
    if not os.environ.get('RAILWAY_ENVIRONMENT'):
        return jsonify({"error": "هذه الوظيفة متاحة فقط على Railway"}), 400

    try:
        # هنا يمكنك إضافة كود لجلب البيانات من مصدر خارجي
        # أو من خلال API من الخادم المحلي
        return jsonify({
            "status": "info",
            "message": "للمزامنة، رفع البيانات يدوياً إلى قاعدة البيانات"
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
def fetch_fingerprint_data(date=None):
    """
    دالة لجلب بيانات البصمة - معدلة لتجنب تعارض الجلسات
    """
    from zk import ZK

    zk = ZK('192.168.1.201', port=4370, timeout=30)
    conn = None
    data = []

    try:
        print("🔌 محاولة الاتصال بجهاز البصمة...")
        conn = zk.connect()
        conn.disable_device()

        print("📥 جلب بيانات الحضور...")
        attendance = conn.get_attendance()
        print(f"📊 عدد السجلات الخام: {len(attendance)}")

        for att in attendance:
            att_time = att.timestamp
            att_date_str = att_time.strftime("%Y-%m-%d")

            # فلترة حسب التاريخ إذا تم تحديده
            if date and att_date_str != date.strftime("%Y-%m-%d"):
                continue

            data.append({
                "employee_code": str(att.user_id),
                "timestamp": att_time
            })

        print(f"✅ تم جلب {len(data)} سجل بعد الفلترة")

    except Exception as e:
        print(f"❌ خطأ في الاتصال بالبصمة: {e}")
        return []  # إرجاع قائمة فارغة بدلاً من البيانات التجريبية

    finally:
        if conn:
            try:
                conn.enable_device()
                conn.disconnect()
                print("🔌 تم فصل الاتصال بالبصمة")
            except:
                pass

    # دمج الدخول والخروج لكل موظف
    return merge_attendance_records(data)


def merge_attendance_records(raw_data):
    """
    دمج سجلات الدخول والخروج لكل موظف
    """
    merged_data = {}

    for rec in raw_data:
        emp_code = rec["employee_code"]
        date_key = rec["timestamp"].strftime("%Y-%m-%d")
        time_key = rec["timestamp"].strftime("%H:%M")

        key = (emp_code, date_key)

        if key not in merged_data:
            merged_data[key] = {
                "employee_code": emp_code,
                "check_in": time_key,
                "check_out": None
            }
        else:
            # تحديث وقت الخروج إذا كان الوقت لاحقاً
            existing_time = merged_data[key]["check_in"]
            if time_key > existing_time:
                merged_data[key]["check_out"] = time_key
            else:
                # إذا كان الوقت سابقاً، جعله دخولاً والخروج الحالي يصبح خروجاً
                merged_data[key]["check_in"] = time_key
                merged_data[key]["check_out"] = existing_time

    result = list(merged_data.values())
    print(f"🔄 تم دمج {len(raw_data)} سجل إلى {len(result)} سجل نهائي")
    return result

# ===================== API لسحب/تحديث البيانات =====================

@bp.route("/sync_fingerprint")
def sync_fingerprint():
    records = fetch_fingerprint_data()
    added_count = 0

    for rec in records:
        employee = Employee.query.filter_by(employee_code=rec["employee_code"]).first()
        if not employee:
            continue

        # استخدم التاريخ من timestamp بدلاً من check_in
        att_date = rec.get("timestamp").date()
        check_in = rec.get("check_in")
        check_out = rec.get("check_out")

        existing = AttendanceRecord.query.filter_by(employee_id=employee.id, date=att_date).first()
        if existing:
            existing.check_in = check_in
            existing.check_out = check_out
        else:
            attendance = AttendanceRecord(
                employee_id=employee.id,
                date=att_date,
                check_in=check_in,
                check_out=check_out
            )
            db.session.add(attendance)
            added_count += 1

    db.session.commit()
    return jsonify({"status": "success", "added": added_count})


#دالة سحب و تحديث بيانات الموظفين الساكنين مهم جدا
# ======================== دالة لجلب بيانات البصمة ========================


@bp.route("/view_combined")
def view_fingerprint_attendance_combined():
    from datetime import datetime
    date_str = request.args.get("date")
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else datetime.today().date()

    fingerprint_records = fetch_fingerprint_data(selected_date)
    combined_records = []

    for rec in fingerprint_records:
        employee = Employee.query.filter_by(employee_code=rec["employee_code"]).first()
        combined_records.append({
            "employee_code": rec["employee_code"],
            "employee": employee,
            "check_in": rec.get("check_in"),
            "check_out": rec.get("check_out")
        })

    return render_template(
        "housing/fingerprint_attendance_combined.html",
        records=combined_records,
        selected_date=selected_date
    )



@bp.route("/view_all")
def fingerprint_attendance_all():
    date_str = request.args.get("date")
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else datetime.today().date()

    # 1️⃣ جلب البيانات مباشرة من الجهاز
    fingerprint_records = fetch_fingerprint_data(selected_date)
    combined_records = []

    for rec in fingerprint_records:
        employee = Employee.query.filter_by(employee_code=rec["employee_code"]).first()
        if not employee:
            continue  # تجاهل الموظفين غير المسجلين في قاعدة البيانات

        # 2️⃣ تحويل النصوص إلى وقت لتوافق SQLite
        check_in_str = rec.get("check_in")
        check_in_time = datetime.strptime(check_in_str, "%H:%M").time() if check_in_str else None

        check_out_str = rec.get("check_out")
        check_out_time = datetime.strptime(check_out_str, "%H:%M").time() if check_out_str else None

        # 3️⃣ حفظ أو تحديث البيانات في قاعدة البيانات
        att_date = selected_date
        existing = AttendanceRecord.query.filter_by(employee_id=employee.id, date=att_date).first()
        if existing:
            if check_in_time:
                existing.check_in = check_in_time
            if check_out_time:
                existing.check_out = check_out_time
        else:
            attendance = AttendanceRecord(
                employee_id=employee.id,
                date=att_date,
                check_in=check_in_time,
                check_out=check_out_time
            )
            db.session.add(attendance)

        # 4️⃣ إضافة البيانات لعرضها مباشرة في الصفحة
        combined_records.append({
            "employee_code": employee.employee_code,
            "name": employee.name,
            "company_name": employee.company.name if employee.company else "-",  # افتراض أن العلاقة موجودة
            "department": employee.department,
            "job_title": employee.job_title,
            "work_type": employee.work_type,
            "check_in": check_in_time.strftime("%H:%M") if check_in_time else "",
            "check_out": check_out_time.strftime("%H:%M") if check_out_time else ""
        })

    db.session.commit()  # حفظ كل التحديثات دفعة واحدة

    return render_template(
        "housing/fingerprint_attendance_all.html",
        records=combined_records,
        selected_date=selected_date
    )

@bp.route("/attendance_sql")
def attendance_sql():
    from datetime import datetime
    from sqlalchemy import text

    date_str = request.args.get("date")
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else datetime.today().date()

    # جلب بيانات الموظفين الذين بصموا فقط
    sql = """
        SELECT 
            e.employee_code,
            e.name,
            e.company_name,
            e.department,
            e.job_title,
            e.work_type,
            MIN(f.check_in) AS check_in,
            MAX(f.check_out) AS check_out
        FROM employees e
        INNER JOIN fingerprint_attendance f 
            ON e.id = f.employee_id
        WHERE DATE(f.date) = :selected_date
        GROUP BY e.employee_code, e.name, e.company_name, e.department, e.job_title, e.work_type
        ORDER BY e.employee_code
    """

    records = db.session.execute(text(sql), {"selected_date": selected_date}).fetchall()

    return render_template(
        "housing/attendance_sql.html",
        records=records,
        selected_date=selected_date
    )


@bp.route("/view")
def view_fingerprint_attendance():
    print("✅ تم استدعاء الدالة بنجاح")

    date_str = request.args.get("date")
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else datetime.today().date()

    print(f"📅 التاريخ المحدد: {selected_date}")

    try:
        # جلب بيانات البصمة
        fingerprint_records = fetch_fingerprint_data(selected_date)
        print(f"📊 عدد سجلات البصمة: {len(fingerprint_records)}")

        combined_records = []

        # استخدام جلسة منفصلة للحذف لتجنب التعارض
        try:
            deleted_count = FingerprintArchive.query.filter_by(date=selected_date).delete()
            db.session.commit()
            print(f"🗑️ تم حذف {deleted_count} سجل قديم")
        except Exception as delete_error:
            print(f"⚠️ لم يتم حذف السجلات القديمة: {delete_error}")
            db.session.rollback()

        # معالجة كل سجل بجلسة منفصلة
        for rec in fingerprint_records:
            try:
                emp_code = str(rec["employee_code"]).strip()

                # البحث عن الموظف باستخدام جلسة منفصلة
                emp_row = None
                try:
                    emp_row = db.session.query(
                        Employee.employee_code,
                        Employee.name,
                        Employee.job_title,
                        Employee.department,
                        Employee.work_type,
                        Employee.company_name,
                        HousingUnit.name.label("housing_name"),
                        Room.room_number
                    ).outerjoin(
                        BedAssignment, Employee.id == BedAssignment.employee_id
                    ).outerjoin(
                        Bed, Bed.id == BedAssignment.bed_id
                    ).outerjoin(
                        Room, Bed.room_id == Room.id
                    ).outerjoin(
                        HousingUnit, Room.housing_unit_id == HousingUnit.id
                    ).filter(
                        Employee.employee_code == emp_code
                    ).first()
                except Exception as query_error:
                    print(f"⚠️ خطأ في استعلام الموظف {emp_code}: {query_error}")
                    continue

                record_data = {
                    "employee_code": emp_code,
                    "name": emp_row.name if emp_row else "غير موجود",
                    "company": emp_row.company_name if emp_row else "-",
                    "job_title": emp_row.job_title if emp_row else "-",
                    "department": emp_row.department if emp_row else "-",
                    "company_housing_name": emp_row.housing_name if emp_row and emp_row.housing_name else "غير مرتبط بسكن",
                    "room_number": emp_row.room_number if emp_row and emp_row.room_number else "-",
                    "work_type": emp_row.work_type if emp_row else "-",
                    "check_in": rec.get("check_in") or "-",
                    "check_out": rec.get("check_out") or "-",
                    "date": selected_date
                }

                # حفظ في الأرشيف بجلسة سريعة
                try:
                    archive_entry = FingerprintArchive(**record_data)
                    db.session.add(archive_entry)
                    # حفظ فوري لكل سجل لتجنب التعارض
                    db.session.commit()
                except Exception as save_error:
                    print(f"⚠️ خطأ في حفظ سجل {emp_code}: {save_error}")
                    db.session.rollback()
                    continue

                combined_records.append(record_data)
                print(f"✅ تم معالجة الموظف {emp_code}")

            except Exception as e:
                print(f"❌ خطأ في معالجة السجل {rec}: {e}")
                db.session.rollback()
                continue

        print(f"🎉 تم معالجة {len(combined_records)} سجل بنجاح")

        return render_template(
            "housing/fingerprint_attendance.html",
            records=combined_records,
            selected_date=selected_date
        )

    except Exception as e:
        print(f"💥 خطأ رئيسي: {e}")
        import traceback
        print(traceback.format_exc())
        db.session.rollback()

        # في حالة الخطأ، نعرض صفحة بدون بيانات بدلاً من بيانات تجريبية
        return render_template(
            "housing/fingerprint_attendance.html",
            records=[],  # قائمة فارغة
            selected_date=selected_date
        ), 500

def get_employee_from_new_db(emp_code):
    return db.session.query(Employee).options(
        joinedload(Employee.company),
        joinedload(Employee.bed).joinedload(Bed.room).joinedload(Room.housing_unit) ).filter(Employee.employee_code == emp_code).first()

from datetime import datetime, timedelta

def archive_fingerprint_data(start_date=None, end_date=None):
    if not start_date: start_date = datetime.today().date()
    if not end_date: end_date = datetime.today().date()
    current_date = start_date
    while current_date <= end_date:
        fingerprint_records = fetch_fingerprint_data(current_date)
        for rec in fingerprint_records:
            try: emp_code = int(float(rec["employee_code"]))
            except: emp_code = rec["employee_code"]
            # جلب الموظف
            emp = Employee.query.filter_by(employee_code=emp_code).first()
            if emp:
                # جلب آخر تخصيص للسرير للموظف في التاريخ الحالي
                assignment = (
                    BedAssignment.query
                    .filter(
                        BedAssignment.employee_id == emp.id,
                        BedAssignment.active == True,
                        BedAssignment.start_date <= current_date,
                        (BedAssignment.end_date == None) | (BedAssignment.end_date >= current_date)
                    )
                    .order_by(BedAssignment.start_date.desc())
                    .first()
                )
                if assignment:
                    bed = assignment.bed
                    room = bed.room
                    housing = room.housing_unit
                    company_housing_name = housing.name
                    room_number = room.room_number
                    bed_number = bed.bed_number
                else:
                    company_housing_name = "فارغ"
                    room_number = "فارغ"
                    bed_number = "فارغ"
                archive_record = FingerprintArchive(
                    employee_code=emp.employee_code,
                    name=emp.name,
                    company=emp.company.name if emp.company else "-",
                    department=emp.department or "-",
                    job_title=emp.job_title or "-",
                    work_type=emp.work_type or "-",
                    company_housing_name=company_housing_name,
                    room_number=room_number,
                    bed_number=bed_number,
                    check_in=rec.get("check_in") or "-",
                    check_out=rec.get("check_out") or "-",
                    date=current_date
                )
            else:
                # موظف غير موجود
                archive_record = FingerprintArchive(
                    employee_code=emp_code,
                    name="غير موجود",
                    company="-",
                    department="-",
                    job_title="-",
                    work_type="-",
                    company_housing_name="فارغ",
                    room_number="فارغ",
                    bed_number="فارغ",
                    check_in=rec.get("check_in") or "-",
                    check_out=rec.get("check_out") or "-",
                    date=current_date
                )
            db.session.add(archive_record)
        db.session.commit()
        current_date += timedelta(days=1)

@bp.route("/archive", methods=["GET", "POST"])
def view_fingerprint_archive():
    # الحصول على القيم من GET أو POST
    date_str = request.args.get("date") or request.form.get("date")
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else None
    search_code = request.args.get("employee_code") or request.form.get("employee_code")
    search_name = request.args.get("name") or request.form.get("name")

    # بدء الاستعلام على أرشيف البصمة
    query = FingerprintArchive.query
    if selected_date: query = query.filter_by(date=selected_date)
    if search_code: query = query.filter(FingerprintArchive.employee_code.contains(search_code))
    if search_name: query = query.filter(FingerprintArchive.name.contains(search_name))

    # ترتيب حسب كود الموظف
    records = query.order_by(FingerprintArchive.employee_code).all()

    # إذا لم يكن هناك أي سجل للسكن/غرفة/سرير، نحاول ملئها مباشرة من BedAssignment
    for r in records:
        if r.company_housing_name in (None, "", "فارغ") or \
           r.room_number in (None, "", "فارغ") or \
           r.bed_number in (None, "", "فارغ"):
            emp = Employee.query.filter_by(employee_code=r.employee_code).first()
            if emp:
                # نأخذ آخر تخصيص للسرير بغض النظر عن التاريخ
                assignment = (
                    BedAssignment.query
                    .filter(BedAssignment.employee_id == emp.id, BedAssignment.active == True)
                    .order_by(BedAssignment.start_date.desc())
                    .first()
                )
                if assignment:
                    bed = assignment.bed
                    room = bed.room
                    housing = room.housing_unit
                    r.company_housing_name = housing.name
                    r.room_number = room.room_number
                    r.bed_number = bed.bed_number

    return render_template(
        "housing/fingerprint_archive.html",
        records=records,
        selected_date=selected_date,
        search_code=search_code or "",
        search_name=search_name or ""
    )

@bp.route("/archive_sync", methods=["GET"])
def archive_fingerprint_sync():
    start_str = request.args.get("start_date")
    end_str = request.args.get("end_date")
    start_date = datetime.strptime(start_str, "%Y-%m-%d").date() if start_str else None
    end_date = datetime.strptime(end_str, "%Y-%m-%d").date() if end_str else None

    # حذف أي أرشيف موجود في الفترة المطلوبة لتجنب التكرار
    if start_date and end_date:
        FingerprintArchive.query.filter(
            FingerprintArchive.date >= start_date,
            FingerprintArchive.date <= end_date
        ).delete()
        db.session.commit()

    # استدعاء الدالة الجديدة للأرشفة
    archive_fingerprint_data(start_date, end_date)

    return "تم أرشفة بيانات البصمة بنجاح!"

# ======================== عرض بيانات البصمة المخزنة ========================
@bp.route("/fingerprint_attendance_stored")
def view_fingerprint_attendance_stored():
    # جلب التاريخ من الباراميتر أو استخدام اليوم الحالي
    date_str = request.args.get("date")
    selected_date = (
        datetime.strptime(date_str, "%Y-%m-%d").date()
        if date_str else datetime.today().date()
    )

    # جلب السجلات من جدول البصمة حسب التاريخ
    records = (
        db.session.query(AttendanceRecord)
        .filter(AttendanceRecord.date == selected_date)
        .all()
    )

    return render_template(
        "housing/fingerprint_attendance_stored.html",
        records=records,
        selected_date=selected_date
    )

from flask import make_response

@bp.route("/export_attendance_excel")
def export_attendance_excel():
    date_str = request.args.get("date")
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else datetime.today().date()

    # جلب بيانات الحضور
    db_records = {r.employee_id: r for r in AttendanceRecord.query.filter_by(date=selected_date).all() if r.employee_id}
    fingerprint_records = fetch_fingerprint_data(selected_date)

    data = []
    for rec in fingerprint_records:
        emp_db = db_records.get(rec.get("employee_id"))
        data.append({
            "كود الموظف": rec.get("employee_id"),
            "الاسم": emp_db.employee.name if emp_db else "غير موجود",
            "تسجيل الدخول": rec.get("check_in"),
            "تسجيل الخروج": rec.get("check_out"),
        })

    df = pd.DataFrame(data)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Attendance")
    output.seek(0)

    response = make_response(output.read())
    response.headers["Content-Disposition"] = f"attachment; filename=attendance_{selected_date}.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


@bp.route("/export_attendance_pdf")
def export_attendance_pdf():
    date_str = request.args.get("date")
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else datetime.today().date()

    db_records = {r.employee_id: r for r in AttendanceRecord.query.filter_by(date=selected_date).all() if r.employee_id}
    fingerprint_records = fetch_fingerprint_data(selected_date)

    pdf = FPDF(orientation='L', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 10, f"تقرير حضور البصمة - {selected_date}", ln=True, align='C')

    pdf.set_font("Arial", "", 11)
    pdf.ln(5)
    col_widths = [40, 80, 40, 40]

    # رأس الجدول
    headers = ["كود الموظف", "الاسم", "تسجيل الدخول", "تسجيل الخروج"]
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header, border=1, align='C')
    pdf.ln()

    # بيانات الجدول
    for rec in fingerprint_records:
        emp_db = db_records.get(rec.get("employee_id"))
        row = [
            str(rec.get("employee_id")),
            emp_db.employee.name if emp_db else "غير موجود",
            str(rec.get("check_in") or "-"),
            str(rec.get("check_out") or "-")
        ]
        for i, val in enumerate(row):
            pdf.cell(col_widths[i], 10, val, border=1, align='C')
        pdf.ln()

    response = make_response(pdf.output(dest='S').encode('latin1'))
    response.headers["Content-Disposition"] = f"attachment; filename=attendance_{selected_date}.pdf"
    response.headers["Content-Type"] = "application/pdf"
    return response

# ======================== تحديث بيانات البصمة ========================
@bp.route("/update_fingerprint")
def update_fingerprint():
    date_str = request.args.get("date")
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else datetime.today().date()

    # سحب البيانات وحفظها مباشرة
    result = fetch_fingerprint_data(selected_date)
    print(result)

    return redirect(url_for("main.view_fingerprint_attendance_db", date=selected_date))

# ======================== عرض بيانات الحضور من قاعدة البيانات ========================

@bp.route("/compare_pdf_with_fingerprint")
def compare_pdf_with_fingerprint():
    from datetime import datetime
    import pdfplumber
    import re

    # 🔹 دالة لتوحيد الأكواد
    def normalize_code(code):
        if not code:
            return ""
        # إزالة المسافات
        code = str(code).strip().replace(" ", "")
        # تحويل الأرقام العربية/الهندية إلى انجليزية
        arabic_to_eng = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
        code = code.translate(arabic_to_eng)
        # إزالة الأصفار البادئة
        code = re.sub(r"^0+", "", code)
        return code

    # 1️⃣ جلب التاريخ
    date_str = request.args.get("date")
    selected_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else datetime.today().date()

    # 2️⃣ بيانات البصمة
    fingerprint_records = fetch_fingerprint_data(selected_date)
    fp_data = []
    for rec in fingerprint_records:
        emp_code = normalize_code(rec.get("employee_code", ""))
        employee = Employee.query.filter_by(employee_code=emp_code).first()
        fp_data.append({
            "employee_code": emp_code,
            "name": employee.name if employee else rec.get("name", "غير موجود"),
            "company": employee.company.name if employee and employee.company else "-",
            "check_in": rec.get("check_in") or "-",
            "check_out": rec.get("check_out") or "-"
        })
    fp_df = pd.DataFrame(fp_data)

    # 3️⃣ قراءة PDF
    pdf_path = r"D:\ghith\NEW\rooms\uploads\eae78a61-7690-4c22-8d95-14349f5bd51d.pdf"
    pdf_employees = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                for row in table[1:]:
                    emp_code = normalize_code(row[0]) if row[0] else ""
                    name = str(row[1]).strip() if len(row) > 1 else ""
                    department = str(row[2]).strip() if len(row) > 2 else ""
                    job_title = str(row[3]).strip() if len(row) > 3 else ""
                    pdf_employees.append({
                        "employee_code": emp_code,
                        "name": name,
                        "department": department,
                        "job_title": job_title
                    })
    pdf_df = pd.DataFrame(pdf_employees)

    # 4️⃣ استخراج الاختلافات بعد التوحيد
    missing_in_fp = pdf_df[~pdf_df["employee_code"].isin(fp_df["employee_code"])]
    missing_in_pdf = fp_df[~fp_df["employee_code"].isin(pdf_df["employee_code"])]

    # ✅ ترتيب حسب الشركة
    if "department" in missing_in_fp:
        missing_in_fp = missing_in_fp.sort_values(by="department", na_position="last")
    if "company" in missing_in_pdf:
        missing_in_pdf = missing_in_pdf.sort_values(by="company", na_position="last")

    # 5️⃣ عرض النتائج
    return render_template(
        "housing/compare_pdf_fingerprint.html",
        selected_date=selected_date,
        missing_in_fp=missing_in_fp.to_dict(orient="records"),
        missing_in_pdf=missing_in_pdf.to_dict(orient="records")
    )

from flask import render_template, jsonify
import json

@bp.route('/housing_map/<int:housing_id>')
def housing_map(housing_id):
    # جلب السكن الحالي
    housing = HousingUnit.query.get_or_404(housing_id)
    # جلب كل السكنات لعرضها كأزرار
    all_housings = HousingUnit.query.order_by(HousingUnit.name).all()

    rooms_data = []
    for room in housing.rooms:
        employees = []
        for bed in room.beds:
            emp = getattr(bed, 'current_employee', None)
            if callable(emp):
                emp = emp()
            if emp:
                employees.append({
                    "name": getattr(emp, "name", ""),
                    "employee_code": getattr(emp, "employee_code", ""),
                    "job_title": getattr(emp, "job_title", "")
                })

        assets = []
        for link in getattr(room, "asset_links", []):
            if link.asset:
                assets.append({
                    "name": link.asset.asset_name,
                    "tag": link.asset.asset_number
                })

        x = getattr(room, "x", None)
        y = getattr(room, "y", None)
        w = getattr(room, "width", None)
        h = getattr(room, "height", None)

        rooms_data.append({
            "id": room.id,
            "room_number": getattr(room, "room_number", ""),
            "total_beds": getattr(room, "total_beds", 0),
            "occupied_beds": getattr(room, "occupied_beds", 0),
            "employees": employees,
            "assets": assets,
            "x": x,
            "y": y,
            "width": w,
            "height": h
        })

    return render_template(
        "housing/housing_map.html",
        housing_name=housing.name,
        rooms_json=json.dumps(rooms_data),
        all_housings=all_housings  # ← مهم جداً
    )


from datetime import datetime, timedelta
from sqlalchemy import func

@bp.route("/housing/report/all")
def housing_report_all():
    # إحصائيات السكنات
    total_housings = HousingUnit.query.count()

    # إحصائيات الموظفين
    total_employees = Employee.query.count()

    # إحصائيات الأسرة
    total_beds = db.session.query(func.sum(HousingUnit.total_beds)).scalar() or 0
    occupied_beds = db.session.query(BedAssignment).filter(
        BedAssignment.active == True,
        BedAssignment.start_date <= datetime.now().date(),
        (BedAssignment.end_date == None) | (BedAssignment.end_date >= datetime.now().date())
    ).count()

    # حساب نسبة الإشغال
    occupancy_rate = round((occupied_beds / total_beds * 100), 1) if total_beds > 0 else 0

    # عدد الغرف الشاغرة
    total_rooms = db.session.query(Room).count()
    occupied_rooms = db.session.query(Room).join(Bed).filter(
        Bed.is_occupied == True
    ).distinct().count()
    vacant_rooms = total_rooms - occupied_rooms

    # عدد التنقلات في الشهر الحالي
    current_month_start = datetime.now().replace(day=1)
    recent_transfers = BedTransfer.query.filter(
        BedTransfer.transfer_date >= current_month_start
    ).count()

    # عدد الشركات
    total_companies = Company.query.count()

    stats = {
        'total_housings': total_housings,
        'total_employees': total_employees,
        'total_beds': total_beds,
        'occupied_beds': occupied_beds,
        'vacant_rooms': vacant_rooms,
        'recent_transfers': recent_transfers,
        'total_companies': total_companies,
        'occupancy_rate': occupancy_rate
    }

    return render_template('housing/housing_report_all.html', stats=stats)

##### الاصـــــــــــــــــــــــــــــــــــــــول
from models import db, Employee


import io

# تحويل النص العربي لعرض صحيح في PDF
def ar(text):
    if not text:
        return ""
    return get_display(arabic_reshaper.reshape(str(text)))


@bp.route('/export/<file_type>')
def export_assets(file_type):
    assets = Asset.query.order_by(Asset.asset_number).all()

    if file_type == 'excel':
        # إنشاء DataFrame
        data = []
        for a in assets:
            data.append({
                "رقم الأصل": a.asset_number,
                "اسم الأصل": a.asset_name,
                "بيان الأصل": a.asset_description,
                "نوع الأصل": a.asset_type,
                "تاريخ الشراء": a.purchase_date.strftime('%Y-%m-%d') if a.purchase_date else '',
                "حالة الأصل": a.status
            })
        df = pd.DataFrame(data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='الأصول')
            writer.save()
        output.seek(0)

        return send_file(output, download_name="assets.xlsx", as_attachment=True)

    elif file_type == 'pdf':
        pdf = FPDF()
        pdf.add_page()
        pdf.add_font("Amiri", "", "D:/ghith/NEW/rooms/fonts/Amiri-Regular.ttf", uni=True)
        pdf.set_font("Amiri", "B", 14)
        pdf.cell(0, 10, ar("تقرير الأصول"), ln=True, align='C')
        pdf.ln(5)

        # جدول PDF
        col_widths = [30, 40, 50, 30, 30, 30]
        headers = ["رقم الأصل", "اسم الأصل", "بيان الأصل", "نوع الأصل", "تاريخ الشراء", "حالة الأصل"]
        line_height = 8

        # ترويسة
        pdf.set_font("Amiri", "B", 12)
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], line_height, ar(h), border=1, align='C')
        pdf.ln()

        # البيانات
        pdf.set_font("Amiri", "", 12)
        for a in assets:
            pdf.cell(col_widths[0], line_height, ar(a.asset_number), border=1, align='C')
            pdf.cell(col_widths[1], line_height, ar(a.asset_name), border=1, align='C')
            pdf.cell(col_widths[2], line_height, ar(a.asset_description), border=1, align='C')
            pdf.cell(col_widths[3], line_height, ar(a.asset_type), border=1, align='C')
            pdf.cell(col_widths[4], line_height, ar(a.purchase_date.strftime('%Y-%m-%d') if a.purchase_date else ''), border=1, align='C')
            pdf.cell(col_widths[5], line_height, ar(a.status), border=1, align='C')
            pdf.ln()

        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)

        return send_file(output, download_name="assets.pdf", as_attachment=True)

    else:
        return "نوع الملف غير مدعوم", 400




# عرض لوحة الأصول

@bp.route('/assets')
def assets_dashboard():
    assets = Asset.query.filter_by(is_active=True).all()  # فقط الأصول الفعالة
    return render_template('assets/assets_list.html', assets=assets)


# إضافة أصل جديد
@bp.route("/add_asset", methods=["GET", "POST"])
def add_asset():
    if request.method == "POST":
        try:
            asset_number = request.form.get("asset_number")

            # تحقق إذا الرقم موجود مسبقًا
            existing_asset = Asset.query.filter_by(asset_number=asset_number).first()
            if existing_asset:
                flash(f"الرقم {asset_number} مستخدم مسبقًا. يرجى اختيار رقم آخر.", "danger")
                return redirect(url_for("main.add_asset"))

            purchase_date_str = request.form.get("purchase_date", "").strip()
            purchase_date = None
            if purchase_date_str:
                purchase_date = datetime.strptime(purchase_date_str, "%Y-%m-%d").date()

            new_asset = Asset(
                asset_number=asset_number,
                asset_name=request.form.get("asset_name"),
                asset_description=request.form.get("asset_description"),
                asset_type=request.form.get("asset_type"),
                purchase_date=purchase_date,
                status=request.form.get("status")
            )

            db.session.add(new_asset)
            db.session.commit()
            flash("تمت إضافة الأصل بنجاح", "success")
            return redirect(url_for("main.add_asset"))

        except Exception as e:
            db.session.rollback()
            flash(f"حدث خطأ أثناء إضافة الأصل: {e}", "danger")

    # جلب كل الأصول لعرضها في الجدول
    all_assets = Asset.query.order_by(Asset.asset_number).all()
    return render_template("assets/add_asset.html", assets=all_assets)

# تعديل أصل موجود

# حذف أصل

@bp.route("/edit_asset/<int:asset_id>", methods=["GET", "POST"])
def edit_asset(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    if request.method == "POST":
        try:
            asset.asset_number = request.form.get("asset_number")
            asset.asset_name = request.form.get("asset_name")
            asset.asset_description = request.form.get("asset_description")
            asset.asset_type = request.form.get("asset_type")

            purchase_date_str = request.form.get("purchase_date", "").strip()
            if purchase_date_str:
                asset.purchase_date = datetime.strptime(purchase_date_str, "%Y-%m-%d").date()
            else:
                asset.purchase_date = None

            asset.status = request.form.get("status")

            db.session.commit()
            flash("تم تعديل الأصل بنجاح", "success")
            return redirect(url_for("main.assets_dashboard"))
        except Exception as e:
            db.session.rollback()
            flash(f"حدث خطأ أثناء تعديل الأصل: {e}", "danger")

    return render_template("assets/edit_asset.html", asset=asset)


# حذف أصل

@bp.route("/delete_asset/<int:asset_id>", methods=["POST"])
def delete_asset(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    try:
        db.session.delete(asset)
        db.session.commit()
        flash("تم حذف الأصل بنجاح", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"حدث خطأ أثناء حذف الأصل: {e}", "danger")
    return redirect(url_for("main.assets_dashboard"))

@bp.route('/link_asset_to_housing', methods=['GET', 'POST'])
@bp.route('/link_asset_to_housing/<int:asset_id>', methods=['GET', 'POST'])
def link_asset_to_housing(asset_id=None):
    # استدعاء جميع البيانات المطلوبة للواجهة
    assets = Asset.query.all()
    housings = HousingUnit.query.all()
    rooms = Room.query.all()
    beds = Bed.query.all()
    asset_links = AssetLink.query.all()

    # إذا جاء asset_id من الرابط، نحضّره ليظهر محددًا في القائمة
    selected_asset = None
    if asset_id:
        selected_asset = Asset.query.get(asset_id)

    if request.method == 'POST':
        form_asset_id = request.form.get('asset_id')
        housing_id = request.form.get('housing_id')
        room_id = request.form.get('room_id') or None
        bed_id = request.form.get('bed_id') or None

        if not form_asset_id or not housing_id:
            flash("يجب اختيار الأصل والسكن على الأقل.", "danger")
            return redirect(url_for('main.link_asset_to_housing'))

        # تحقق من عدم تكرار الربط
        existing = AssetLink.query.filter_by(
            asset_id=form_asset_id,
            housing_unit_id=housing_id,
            room_id=room_id,
            bed_id=bed_id
        ).first()

        if existing:
            flash("هذا الربط موجود مسبقًا.", "warning")
        else:
            existing_any = AssetLink.query.filter_by(asset_id=form_asset_id).first()
            if existing_any:
                flash("هذا الأصل مرتبط مسبقًا بأي مكان، لا يمكن ربطه مرة أخرى.", "warning")
            else:
                link = AssetLink(
                    asset_id=form_asset_id,
                    housing_unit_id=housing_id,
                    room_id=room_id,
                    bed_id=bed_id,
                    link_date=datetime.utcnow().date()
                )
                db.session.add(link)
                db.session.commit()
                flash("تم ربط الأصل بالسكن/الغرفة/السرير بنجاح", "success")

        return redirect(url_for('main.link_asset_to_housing'))

    return render_template(
        "assets/link_asset_to_housing.html",
        assets=assets,
        housings=housings,
        rooms=rooms,
        beds=beds,
        asset_links=asset_links,
        selected_asset=selected_asset  # هذا سيجعل الأصل يظهر مباشرة في القائمة
    )

@bp.route("/api/rooms/<int:housing_id>")
def api_rooms(housing_id):
    rooms = Room.query.filter_by(housing_unit_id=housing_id).all()
    rooms_data = [{"id": r.id, "room_number": r.room_number} for r in rooms]
    return jsonify(rooms_data)

from models import  Bed

@bp.route('/edit_asset_link/<int:link_id>', methods=['GET', 'POST'])
def edit_asset_link(link_id):
    link = AssetLink.query.get_or_404(link_id)
    assets = Asset.query.all()
    housings = HousingUnit.query.all()
    rooms = Room.query.all()
    beds = Bed.query.all()
    action_type = 'تعديل'

    if request.method == 'POST':
        asset_id = request.form['asset_id']
        housing_id = request.form['housing_id']
        room_id = request.form.get('room_id') or None
        bed_id = request.form.get('bed_id') or None

        # تحقق إذا كان هناك تغيير فعلي
        if (link.asset_id != int(asset_id) or
            link.housing_unit_id != int(housing_id) or
            (link.room_id != int(room_id) if room_id else link.room_id is not None) or
            (link.bed_id != int(bed_id) if bed_id else link.bed_id is not None)):

            # معلومات النقل القديمة والجديدة
            old_housing = link.housing_unit.name if link.housing_unit else "بدون سكن"
            new_housing = HousingUnit.query.get(housing_id).name

            old_room = link.room.room_number if link.room else " -----"
            new_room = Room.query.get(room_id).room_number if room_id else " ----- "

            old_bed = link.bed.bed_number if link.bed else " ----- "
            new_bed = Bed.query.get(bed_id).bed_number if bed_id else " ----- "

            description = f'تم تعديل الربط: السكن {old_housing} → {new_housing}, الغرفة {old_room} → {new_room}, السرير {old_bed} → {new_bed}'

            # تسجيل الإجراء في AssetAction
            new_action = AssetAction(
                asset_id=link.asset_id,
                action_type=action_type,
                description=description,
                action_date=datetime.utcnow().date(),
                old_housing=old_housing,
                old_room=old_room,
                old_bed=old_bed,
                new_housing=new_housing,
                new_room=new_room,
                new_bed=new_bed,
                purchase_date=link.asset.purchase_date,
                disposal_date=link.asset.disposal[-1].disposal_date if link.asset.disposal else None
            )

            db.session.add(new_action)

            # تحديث الربط فعليًا في AssetLink
            link.asset_id = asset_id
            link.housing_unit_id = housing_id
            link.room_id = room_id
            link.bed_id = bed_id
            link.link_date = datetime.utcnow().date()  # تحديث تاريخ الربط

            db.session.commit()
            flash('تم تعديل ربط الأصل بنجاح!', 'success')
        else:
            flash('لم يتم تعديل أي بيانات.', 'info')

        return redirect(url_for('main.link_asset_to_housing'))

    return render_template(
        'assets/edit_asset_link.html',
        link=link,
        assets=assets,
        housings=housings,
        rooms=rooms,
        beds=beds
    )


from flask import request, redirect, url_for, flash

@bp.route("/dispose_asset/<int:asset_id>", methods=["GET", "POST"])
def dispose_asset(asset_id):
    asset = Asset.query.get_or_404(asset_id)

    if request.method == "POST":
        reason = request.form["reason"]

        # تسجيل الإجراء في AssetAction
        old_housing = asset.links[-1].housing_unit.name if asset.links else None
        old_room = asset.links[-1].room.room_number if asset.links and asset.links[-1].room else None
        old_bed = asset.links[-1].bed.bed_number if asset.links and asset.links[-1].bed else None

        new_action = AssetAction(
            asset_id=asset.id,
            action_type="اتلاف",
            description=f"تم اتلاف الأصل. السبب: {reason}",
            action_date=datetime.utcnow().date(),
            old_housing=old_housing,
            old_room=old_room,
            old_bed=old_bed,
            new_housing=None,
            new_room=None,
            new_bed=None,
            purchase_date=asset.purchase_date,
            disposal_date=datetime.utcnow().date()
        )
        db.session.add(new_action)

        # استبعاد الأصل من قائمة الأصول الفعالة
        asset.is_active = False
        asset.status = "ملحق اتلاف"

        # حذف أي روابط سكن حالية
        AssetLink.query.filter_by(asset_id=asset.id).delete()

        # إضافة سجل التصرف النهائي
        disposal = AssetDisposal(
            asset_id=asset.id,
            disposal_reason=reason,
            disposal_date=datetime.utcnow().date()
        )
        db.session.add(disposal)

        db.session.commit()
        flash("تم تسجيل اتلاف الأصل بنجاح!", "success")
        return redirect(url_for("main.assets_dashboard"))  # قائمة الأصول المتاحة

    return render_template("assets/dispose_asset.html", asset=asset)


# واجهة Dashboard لتقارير الأصول
@bp.route("/assets_reports")
def assets_reports_dashboard():
    return render_template("assets/assets_reports_dashboard.html")


@bp.route("/reports/details")
def assets_report():
    assets = Asset.query.order_by(Asset.asset_number).all()

    # حساب عمر الأصل لكل أصل
    now = datetime.now()
    for asset in assets:
        if asset.purchase_date:
            delta = now - asset.purchase_date
            asset.asset_age = round(delta.days / 365, 1)
        else:
            asset.asset_age = None

    # إنشاء قاموس: لكل بيان أصل → لكل اسم أصل → عدد الأصول
    assets_summary = {}
    # إجمالي لكل بيان أصل
    assets_total_by_description = {}

    for asset in assets:
        description = getattr(asset, 'asset_description', 'غير محدد')
        name = getattr(asset, 'asset_name', 'غير محدد')

        if description not in assets_summary:
            assets_summary[description] = {}
        assets_summary[description][name] = assets_summary[description].get(name, 0) + 1

        # حساب الإجمالي حسب البيان
        assets_total_by_description[description] = assets_total_by_description.get(description, 0) + 1
    # إضافة تجميع جديد: إجمالي لكل اسم أصل
    assets_total_by_name = {}
    for asset in assets:
        name = getattr(asset, 'asset_name', 'غير محدد')
        assets_total_by_name[name] = assets_total_by_name.get(name, 0) + 1

    return render_template(
        "assets/assets_report.html",
        assets=assets,
        assets_summary=assets_summary,
        assets_total_by_description=assets_total_by_description,
        assets_total_by_name=assets_total_by_name  # ✅ تمرير البيانات الجديدة
    )


@bp.route("/reports/linked_assets")
def linked_assets_report():
    asset_links = AssetLink.query \
        .join(Asset, Asset.id == AssetLink.asset_id) \
        .join(AssetLink.housing_unit) \
        .outerjoin(AssetLink.room) \
        .outerjoin(AssetLink.bed) \
        .add_entity(Asset) \
        .all()

    data = []
    for link, asset in asset_links:
        if asset.purchase_date:
            purchase_date = asset.purchase_date.date() if hasattr(asset.purchase_date, 'date') else asset.purchase_date
            age_years = round((date.today() - purchase_date).days / 365.25, 1)
        else:
            age_years = 0

        data.append({
            "asset_link": link,
            "asset": asset,
            "age": age_years,
            "housing": link.housing_unit.name if link.housing_unit else "بدون سكن",
            "room": link.room.room_number if link.room else "-----",
            "bed": link.bed.bed_number if link.bed else "------ "
        })

    # خلاصة لكل وحدة سكنية وعدد كل نوع أصل
    summary = {}
    for item in data:
        housing = item["housing"]
        asset_name = item["asset"].asset_name
        summary.setdefault(housing, {})
        summary[housing][asset_name] = summary[housing].get(asset_name, 0) + 1

    # تحليل ومقترحات بالذكاء الاصطناعي
    ai_analysis = []
    for housing, assets_count in summary.items():
        for asset_name, count in assets_count.items():
            if count > 5:
                ai_analysis.append(f"الوحدة {housing}: عدد {asset_name} كبير، يُفضل مراجعة التخزين.")
            elif count == 1:
                ai_analysis.append(f"الوحدة {housing}: {asset_name} فردي، يُنصح بالتحقق من الحاجة.")

    return render_template(
        "assets/linked_assets_report.html",
        asset_links=data,
        summary=summary,
        ai_analysis=ai_analysis
    )


@bp.route("/reports/rooms_asset")
def report_rooms_assets():
    try:
        # جلب البيانات مع استعلام أكثر كفاءة
        asset_links = AssetLink.query \
            .options(
            joinedload(AssetLink.asset),
            joinedload(AssetLink.housing_unit),
            joinedload(AssetLink.room),
            joinedload(AssetLink.bed)
        ) \
            .join(Asset, Asset.id == AssetLink.asset_id) \
            .all()

        data = []
        today = date.today()

        for link in asset_links:
            asset = link.asset

            # حساب عمر الأصل
            age_years = 0
            if asset.purchase_date:
                purchase_date = asset.purchase_date.date() if hasattr(asset.purchase_date,
                                                                      'date') else asset.purchase_date
                try:
                    age_days = (today - purchase_date).days
                    age_years = round(age_days / 365.25, 1) if age_days > 0 else 0
                except (TypeError, ValueError):
                    age_years = 0

            data.append({
                "asset_link": link,
                "asset": asset,
                "age": age_years,
                "housing": link.housing_unit.name if link.housing_unit else "بدون سكن",
                "room": link.room.room_number if link.room else "بدون غرفة",
                "bed": link.bed.bed_number if link.bed else "بدون سرير"
            })

        # 🔹 إعادة ترتيب البيانات لتصبح (الوحدة → الغرفة → الأصول)
        report_data = {}
        for item in data:
            housing = item["housing"]
            room = item["room"]

            if housing not in report_data:
                report_data[housing] = {}

            if room not in report_data[housing]:
                report_data[housing][room] = []

            report_data[housing][room].append(item)

        # 🔹 خلاصة لكل وحدة سكنية
        summary = {}
        total_assets_count = 0
        total_housings = len(report_data)
        total_rooms = 0

        for housing, rooms in report_data.items():
            summary[housing] = {"total": 0, "by_type": {}}
            total_rooms += len(rooms)

            for room, assets in rooms.items():
                assets_count = len(assets)
                summary[housing]["total"] += assets_count
                total_assets_count += assets_count

                for item in assets:
                    asset_name = item["asset"].asset_name
                    summary[housing]["by_type"][asset_name] = summary[housing]["by_type"].get(asset_name, 0) + 1

        # 🔹 تحليل ومقترحات بالذكاء الاصطناعي
        ai_analysis = []

        # تحليل التوزيع
        if total_assets_count > 0:
            for housing, details in summary.items():
                housing_percentage = (details["total"] / total_assets_count) * 100

                if housing_percentage > 50:
                    ai_analysis.append(
                        f"الوحدة {housing}: تحتوي على {housing_percentage:.1f}% من إجمالي الأصول، قد تحتاج لإعادة توزيع.")
                elif housing_percentage < 10:
                    ai_analysis.append(f"الوحدة {housing}: تحتوي على نسبة قليلة ({housing_percentage:.1f}%) من الأصول.")

                # تحليل أنواع الأصول
                for asset_name, count in details["by_type"].items():
                    if count > 5:
                        ai_analysis.append(
                            f"الوحدة {housing}: عدد {asset_name} كبير ({count} قطعة)، يُفضل مراجعة التخزين.")
                    elif count == 1:
                        ai_analysis.append(f"الوحدة {housing}: {asset_name} فردي، يُنصح بالتحقق من الحاجة.")

            # تحليل عام
            avg_assets_per_housing = total_assets_count / total_housings if total_housings > 0 else 0
            if avg_assets_per_housing > 10:
                ai_analysis.append(
                    f"متوسط الأصول لكل وحدة سكنية مرتفع ({avg_assets_per_housing:.1f})، قد تحتاج لإعادة توزيع.")

        # إذا لم توجد بيانات
        if not data:
            ai_analysis.append("لا توجد أصول مرتبطة حالياً. يُنصح بربط الأصول بالسكن والغرف.")

        # 🔹 إحصائيات إضافية للقالب
        housing_list = list(report_data.keys())

        # حساب عدد الأصول غير المرتبطة (إذا كان ذلك ممكناً)
        total_assets_in_db = Asset.query.count()
        unlinked_assets_count = total_assets_in_db - total_assets_count

        return render_template(
            "assets/report_rooms_assets.html",
            report_data=report_data,
            summary=summary,
            ai_analysis=ai_analysis,
            housings=housing_list,
            # الإحصائيات المحسوبة مسبقاً
            total_assets_count=total_assets_count,
            total_housings=total_housings,
            total_rooms=total_rooms,
            unlinked_assets_count=unlinked_assets_count,
            # للتوافق مع الإصدار السابق
            totals=total_assets_count
        )

    except Exception as e:
        # معالجة الأخطاء
        current_app.logger.error(f"Error in report_rooms_assets: {str(e)}")

        # إرجاع بيانات فارغة في حالة الخطأ
        return render_template(
            "assets/report_rooms_assets.html",
            report_data={},
            summary={},
            ai_analysis=[f"حدث خطأ في تحميل التقرير: {str(e)}"],
            housings=[],
            total_assets_count=0,
            total_housings=0,
            total_rooms=0,
            unlinked_assets_count=0,
            totals=0
        )
from flask import Response

# ======================
#  تصدير Excel
# ======================
@bp.route("/export/linked_assets_excel")
def export_linked_assets_excel():
    asset_links = AssetLink.query.all()

    # ترتيب الغرف تنازليًا
    asset_links_sorted = sorted(
        asset_links,
        key=lambda x: x.room.room_number if x.room else "",
        reverse=True
    )

    data = []
    for link in asset_links_sorted:
        purchase_date = link.asset.purchase_date.date() if hasattr(link.asset.purchase_date, "date") else link.asset.purchase_date
        age_years = round((date.today() - purchase_date).days / 365.25, 1) if purchase_date else 0

        data.append({
            "رقم الأصل": link.asset.asset_number,
            "اسم الأصل": link.asset.asset_name,
            "نوع الأصل": link.asset.asset_type,
            "حالة الأصل": link.asset.status,
            "السكن الحالي": link.housing_unit.name if link.housing_unit else "بدون سكن",
            "الغرفة الحالية": link.room.room_number if link.room else "بدون غرفة",
            "السرير الحالي": link.bed.bed_number if link.bed else "بدون سرير",
            "تاريخ آخر نقل/ربط": link.link_date.strftime("%Y-%m-%d"),
            "تاريخ الشراء": purchase_date.strftime("%Y-%m-%d") if purchase_date else "",
            "عمر الأصل (سنوات)": age_years
        })

    df = pd.DataFrame(data)

    # استخدام BytesIO لتجهيز الملف للتحميل
    from io import BytesIO
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="الأصول المرتبطة")

        # تنسيق إضافي: عرض الأعمدة
        worksheet = writer.sheets["الأصول المرتبطة"]
        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=linked_assets.xlsx"}
    )

# ======================
#  تصدير PDF
# ======================
from io import BytesIO
from flask import send_file, Blueprint
from fpdf import FPDF
import arabic_reshaper
from bidi.algorithm import get_display
from datetime import date, datetime
from collections import Counter


class PDF(FPDF):
    logo_path = r"D:\ghith\NEW\rooms\static\images\logo.png"
    system_name = "نظام إدارة الأصول والسكن"
    designer_name = "تم تصميمه بواسطة: علي  مبارك"
    company_name = "الشركة اليمنية لتكرير السكر"
    dept_name = "إدارة الشؤون الإدارية - قسم السكنات"

    def header(self):
        # تعريف الخطوط
        self.add_font("Amiri", "", r"D:\ghith\NEW\rooms\fonts\Amiri-Regular.ttf", uni=True)
        self.add_font("Amiri", "B", r"D:\ghith\NEW\rooms\fonts\Amiri-Bold.ttf", uni=True)

        # الشعار (يمين أعلى)
        if self.logo_path:
            self.image(self.logo_path, 5, 8, 7)

        # اسم النظام (وسط)
        self.set_xy(0, 10)
        self.set_font("Amiri", "B", 14)
        self.cell(0, 10, get_display(arabic_reshaper.reshape(self.system_name)), ln=True, align="C")

        # الشركة / الإدارة / التاريخ (يمين أعلى)
        self.set_xy(0, 5)
        self.set_font("Amiri", "", 8)
        company_text = get_display(arabic_reshaper.reshape(self.company_name))
        dept_text = get_display(arabic_reshaper.reshape(self.dept_name))
        today_text = get_display(arabic_reshaper.reshape(f"التاريخ: {date.today().strftime('%Y-%m-%d')}"))
        self.cell(0, 4, f"{company_text} ", ln=True, align="R")
        self.cell(0, 4, f"{dept_text}", ln=True, align="R")
        self.cell(0, 4, today_text, ln=True, align="R")
        self.ln(3)

        # رؤوس الأعمدة (من اليمين لليسار)
        self.set_font("Amiri", "B", 10)
        self.headers = ["رقم الأصل", "اسم الأصل", "نوع", "حالة", "سكن", "غرفة", "سرير", "تاريخ شراء", "عمر الاصل"]
        page_width = 190
        self.col_widths = [page_width * 0.08, page_width * 0.18, page_width * 0.12, page_width * 0.1,
                           page_width * 0.08, page_width * 0.1, page_width * 0.1, page_width * 0.16, page_width * 0.08]
        # إعداد نسخ RTL لاستخدامها لاحقاً في الصفوف
        self.headers_rtl = list(reversed(self.headers))
        self.col_widths_rtl = list(reversed(self.col_widths))

        self.set_fill_color(200, 200, 200)
        for i in range(len(self.headers_rtl)):
            self.cell(self.col_widths_rtl[i], 8,
                      get_display(arabic_reshaper.reshape(self.headers_rtl[i])),
                      1, 0, "C", fill=True)
        self.ln()

    def footer(self):
        self.set_y(-20)
        self.set_font("Amiri", "", 8)
        self.cell(0, 5, get_display(arabic_reshaper.reshape(self.designer_name)), ln=True, align="C")
        self.cell(0, 5, get_display(arabic_reshaper.reshape(f"صفحة {self.page_no()}")), align="C")


@bp.route("/export/linked_assets_pdf")
def export_linked_assets_pdf():
    asset_links_sorted = sorted(
        AssetLink.query.all(),
        key=lambda x: (
            x.housing_unit.name if x.housing_unit else "",
            x.asset.asset_type,
            x.room.room_number if x.room else ""
        )
    )

    pdf = PDF()
    # تأكد أن الخطوط معرّفة قبل أول set_font
    pdf.add_font("Amiri", "", r"D:\ghith\NEW\rooms\fonts\Amiri-Regular.ttf", uni=True)
    pdf.add_font("Amiri", "B", r"D:\ghith\NEW\rooms\fonts\Amiri-Bold.ttf", uni=True)
    pdf.set_font("Amiri", "", 9)
    pdf.add_page()  # سيرسم الهيدر والرؤوس ويضبط col_widths_rtl

    last_housing = None
    housing_assets = []
    fill = False  # تناوب الألوان

    for idx, link in enumerate(asset_links_sorted):
        current_housing = link.housing_unit.name if link.housing_unit else "بدون سكن"

        # بداية سكن جديد
        if last_housing != current_housing:
            # خلاصة السكن السابق (إن وجد)
            if last_housing is not None and housing_assets:
                pdf.ln(3)
                pdf.set_font("Amiri", "B", 9)
                pdf.cell(0, 8,
                         get_display(arabic_reshaper.reshape(f"خلاصة الأصول في السكن: {last_housing}")),
                         ln=True, align="C")
                type_counter = Counter([asset.asset_name for asset in housing_assets])
                for asset_name, count in type_counter.items():
                    summary_text = f"{asset_name}: {count} قطعة"
                    pdf.cell(0, 6, get_display(arabic_reshaper.reshape(summary_text)), ln=True, align="R")
                pdf.ln(5)
                housing_assets = []

                # صفحة جديدة لكل سكن بعد الأول (الرؤوس ستطبع تلقائياً من header)
                pdf.add_page()
                fill = False

            # عنوان السكن
            pdf.set_font("Amiri", "B", 10)
            pdf.cell(0, 8,
                     get_display(arabic_reshaper.reshape(f"السكن: {current_housing}")),
                     ln=True, align="C")
            pdf.set_font("Amiri", "", 9)

        # لتلخيص السكن
        housing_assets.append(link.asset)

        # حساب العمر
        purchase_date = link.asset.purchase_date
        if purchase_date:
            if isinstance(purchase_date, date) and not isinstance(purchase_date, datetime):
                purchase_date_dt = datetime.combine(purchase_date, datetime.min.time())
            else:
                purchase_date_dt = purchase_date
            age_years = round((date.today() - purchase_date_dt.date()).days / 365.25, 1)
        else:
            age_years = 0

        # الصف بالترتيب الأصلي (يسار→يمين منطقياً)
        row = [
            str(link.asset.asset_number),                           # رقم الأصل (رقم)
            link.asset.asset_name,                                  # اسم الأصل (عربي)
            link.asset.asset_type,                                  # نوع (عربي)
            link.asset.status,                                      # حالة (عربي)
            current_housing,                                        # سكن (عربي)
            link.room.room_number if link.room else " ----",        # غرفة (رمز/رقم)
            link.bed.bed_number if link.bed else " ----",           # سرير (رمز/رقم)
            purchase_date.strftime("%Y-%m-%d") if purchase_date else "",  # تاريخ شراء
            str(age_years)                                          # عمر الأصل
        ]

        # نعكس الصف ليتوافق مع الأعمدة المطبوعة RTL
        row_rtl = list(reversed(row))

        # تناوب لون الصفوف
        pdf.set_fill_color(240, 240, 240) if fill else pdf.set_fill_color(255, 255, 255)
        fill = not fill

        # الأعمدة التي تحتاج RTL: (كل ما يحتوي نص عربي)
        # asset_name, asset_type, status, housing, room_number
        for i, val in enumerate(row_rtl):
            if i in (3, 4, 5, 6, 7):  # لاحظ أضفت العمودين تبع السكن والغرفة
                val = get_display(arabic_reshaper.reshape(str(val)))
            pdf.cell(pdf.col_widths_rtl[i], 8, str(val), 1, 0, "C", fill=True)

        pdf.ln()

        last_housing = current_housing

    # خلاصة آخر سكن
    if housing_assets:
        pdf.ln(3)
        pdf.set_font("Amiri", "B", 9)
        pdf.cell(0, 8,
                 get_display(arabic_reshaper.reshape(f"خلاصة الأصول في السكن: {last_housing}")),
                 ln=True, align="C")
        type_counter = Counter([asset.asset_name for asset in housing_assets])
        for asset_name, count in type_counter.items():
            summary_text = f"{asset_name}: {count} قطعة"
            pdf.cell(0, 6, get_display(arabic_reshaper.reshape(summary_text)), ln=True, align="R")

    # إخراج PDF
    pdf_stream = BytesIO()
    pdf.output(pdf_stream)
    pdf_stream.seek(0)
    return send_file(
        pdf_stream,
        mimetype="application/pdf",
        download_name="linked_assets.pdf",
        as_attachment=True
    )

from models import AssetAction, AssetLink, Asset

@bp.route("/reports/movements")
def assets_movements_report():
    # فلترة حسب نوع العملية، افتراضيًا "جميع العمليات"
    action_type_filter = request.args.get("action_type", "all")
    query = AssetAction.query.join(Asset)
    if action_type_filter != "all":
        query = query.filter(AssetAction.action_type == action_type_filter)

    actions = query.order_by(AssetAction.action_date.desc()).all()
    return render_template("assets/assets_movements_report.html", actions=actions, action_type_filter=action_type_filter)

from flask import render_template, request
from sqlalchemy import literal, or_
from models import db, Asset, AssetDisposal, AssetLink, HousingUnit, Room
from sqlalchemy.orm import aliased
from datetime import datetime

def format_date(date_value):
    """ترجع التاريخ بصيغة YYYY-MM-DD أو نص فارغ"""
    if not date_value:
        return ""
    # إذا كان التاريخ نص
    if isinstance(date_value, str):
        try:
            date_obj = datetime.fromisoformat(date_value)
            return date_obj.strftime("%Y-%m-%d")
        except ValueError:
            return date_value  # لو النص مش قابل للتحويل، نرجعه زي ما هو
    # إذا كان datetime
    return date_value.strftime("%Y-%m-%d")

@bp.route('/assets_history', methods=['GET', 'POST'])
def assets_history():
    # فلاتر من الطلب
    asset_id = request.args.get("asset_id")
    action_type_filter = request.args.get("action_type")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # استعلام عمليات النقل (AssetLink)
    links_query = db.session.query(
        Asset.id.label("asset_id"),
        Asset.asset_name.label("asset_name"),
        literal("نقل").label("action_type"),
        AssetLink.link_date.label("action_date"),
        AssetLink.housing_unit_id.label("old_housing_id"),
        AssetLink.room_id.label("old_room_id"),
        AssetLink.bed_id.label("old_bed_id"),
        literal(None).label("description"),
        literal(None).label("purchase_date"),
        literal(None).label("disposal_date")
    ).join(Asset, Asset.id == AssetLink.asset_id)

    # استعلام عمليات الإتلاف (AssetDisposal)
    disposals_query = db.session.query(
        Asset.id.label("asset_id"),
        Asset.asset_name.label("asset_name"),
        literal("اتلاف").label("action_type"),
        AssetDisposal.disposal_date.label("action_date"),
        literal(None).label("old_housing_id"),
        literal(None).label("old_room_id"),
        literal(None).label("old_bed_id"),
        AssetDisposal.disposal_reason.label("description"),
        Asset.purchase_date.label("purchase_date"),
        AssetDisposal.disposal_date.label("disposal_date")
    ).join(Asset, Asset.id == AssetDisposal.asset_id)

    # دمج الاستعلامات
    history_subquery = links_query.union_all(disposals_query).subquery()
    history_query = db.session.query(history_subquery)

    # تطبيق الفلاتر
    if asset_id:
        history_query = history_query.filter(history_subquery.c.asset_id == int(asset_id))
    if action_type_filter:
        history_query = history_query.filter(history_subquery.c.action_type == action_type_filter)
    if start_date:
        history_query = history_query.filter(history_subquery.c.action_date >= start_date)
    if end_date:
        history_query = history_query.filter(history_subquery.c.action_date <= end_date)

    # ترتيب حسب التاريخ تنازلي
    rows = history_query.order_by(literal_column("action_date").desc()).all()

    # جلب أسماء الوحدات/الغرف/الأسرة للعرض بدل الـ id
    housing_map = {h.id: h.name for h in HousingUnit.query.all()}
    room_map = {r.id: r.room_number for r in Room.query.all()}
    bed_map = {b.id: b.bed_number for b in Bed.query.all()}

    # نحول كل Row إلى dict قابل للتعديل
    history = []
    for row in rows:
        history.append({
            "asset_id": row.asset_id,
            "asset_name": row.asset_name,
            "action_type": row.action_type,
            "action_date": format_date(row.action_date),
            "old_housing": housing_map.get(row.old_housing_id),
            "old_room": room_map.get(row.old_room_id),
            "old_bed": bed_map.get(row.old_bed_id),
            "description": row.description,
            "purchase_date": format_date(row.purchase_date),
            "disposal_date": format_date(row.disposal_date),
        })

    # نرجع للواجهة
    return render_template(
        "assets/assets_history.html",
        history=history,
        selected_action_type=action_type_filter
    )

# داخل routes.py
@bp.route('/evaluation_dashboard')
def evaluation_dashboard():
    return redirect("http://192.168.6.61:8000/login")



# الصفحة الرئيسية للـ Blueprint → توجه مباشرة لنموذج الطلب الجديد
@bp.route('/')
def index():
    return redirect(url_for('maintenance_dashboard'))

# لوحة الصيانة: عرض كل الطلبات
@bp.route('/dashboard')
def maintenance_dashboard():
    requests = MaintenanceRequest.query.order_by(MaintenanceRequest.scheduled_date.desc()).all()
    return render_template('maintenance/maintenance_dashboard.html', requests=requests)


# إضافة طلب جديد

from models import db, MaintenanceRequest, HousingUnit, Room, MaintenanceTeam
@bp.route("/maintenance/new", methods=["GET", "POST"])
def new_request():
    housings = HousingUnit.query.options(db.joinedload(HousingUnit.rooms)).all()
    teams = MaintenanceTeam.query.all()  # الفرق موجودة للعرض في قائمة منسدلة

    if request.method == "POST":
        housing_id = request.form.get("housing_id", "").strip()
        room_id = request.form.get("room_id", "").strip()
        title = request.form.get("title", "").strip()
        team = request.form.get("team", "").strip()
        responsible = request.form.get("responsible", "").strip()
        scheduled_date = request.form.get("scheduled_date", "").strip()
        duration_hours = request.form.get("duration_hours", "").strip()
        priority = request.form.get("priority", "").strip()
        maintenance_type = request.form.get("maintenance_type", "").strip()
        manufacturing_order = request.form.get("manufacturing_order", "").strip()
        notes = request.form.get("notes", "").strip()

        # التحقق من الحقول المطلوبة
        if not housing_id or not room_id or not title or not responsible or not scheduled_date:
            flash("يرجى ملء جميع الحقول المطلوبة", "danger")
            return redirect(url_for("main.new_request"))

        housing = HousingUnit.query.get(housing_id)
        room = Room.query.get(room_id)
        if not housing or not room:
            flash("السكن أو الغرفة غير موجودة", "danger")
            return redirect(url_for("main.new_request"))

        try:
            scheduled_date_obj = datetime.strptime(scheduled_date, "%Y-%m-%d").date()
        except ValueError:
            flash("صيغة التاريخ غير صحيحة", "danger")
            return redirect(url_for("main.new_request"))

        try:
            duration_hours_val = float(duration_hours) if duration_hours else 0
        except ValueError:
            flash("مدة الساعات غير صحيحة", "danger")
            return redirect(url_for("main.new_request"))

        # حفظ الطلب
        new_req = MaintenanceRequest(
            title=title,
            team=team,
            responsible=responsible,
            scheduled_date=scheduled_date_obj,
            duration_hours=duration_hours_val,
            priority=priority,
            maintenance_type=maintenance_type,
            manufacturing_order=manufacturing_order,
            notes=notes,
            housing_id=housing.id,
            room_id=room.id
        )

        db.session.add(new_req)
        db.session.commit()
        flash(f"تم حفظ طلب الصيانة '{title}' بنجاح", "success")
        return redirect(url_for("main.maintenance_dashboard"))

    return render_template(
        "maintenance/new_request.html",
        housings=housings,
        teams=teams,
        req=None
    )

# تعديل طلب موجود
@bp.route('/edit/<int:request_id>', methods=['GET', 'POST'])
def edit_request(request_id):
    req = MaintenanceRequest.query.get_or_404(request_id)
    if request.method == 'POST':
        req.title = request.form.get('title')
        req.team = request.form.get('team')
        req.responsible = request.form.get('responsible')
        req.scheduled_date = request.form.get('scheduled_date')
        req.duration_hours = float(request.form.get('duration_hours') or 0)
        req.priority = request.form.get('priority')
        req.maintenance_type = request.form.get('maintenance_type')
        req.manufacturing_order = request.form.get('manufacturing_order')
        req.notes = request.form.get('notes')
        db.session.commit()
        flash("تم تعديل الطلب بنجاح!", "success")
        return redirect(url_for('maintenance_dashboard'))

    return render_template('maintenance/new_request.html', req=req)

# حذف طلب
@bp.route('/delete/<int:request_id>')
def delete_request(request_id):
    req = MaintenanceRequest.query.get_or_404(request_id)
    db.session.delete(req)
    db.session.commit()
    flash("تم حذف الطلب بنجاح!", "success")
    return redirect(url_for('maintenance_dashboard'))


from flask import Blueprint, render_template, request, redirect, url_for, flash
from models import db, MaintenanceTeam


@bp.route("/maintenance_teams", methods=["GET", "POST"])
def maintenance_teams():
    # إذا تم إرسال الفورم من أجل إضافة فريق
    if request.method == "POST":
        name = request.form.get("name")
        if name:
            team = MaintenanceTeam(name=name)
            db.session.add(team)
            db.session.commit()
            flash(f"تم إضافة الفريق '{name}' بنجاح", "success")
        return redirect(url_for("main.maintenance_teams"))

    teams = MaintenanceTeam.query.all()
    return render_template("maintenance/maintenance_teams.html", teams=teams)

@bp.route("/maintenance_teams/delete/<int:team_id>")
def maintenance_team_delete(team_id):
    team = MaintenanceTeam.query.get_or_404(team_id)
    db.session.delete(team)
    db.session.commit()
    flash(f"تم حذف الفريق '{team.name}' بنجاح", "success")
    return redirect(url_for("main.maintenance_teams"))

@bp.route("/maintenance_teams/add", methods=["POST"])
def maintenance_team_add():
    name = request.form.get("name")
    if name:
        team = MaintenanceTeam(name=name)
        db.session.add(team)
        db.session.commit()
    return redirect(url_for("main.maintenance_teams"))


from models import Assetwarehouse, Consumable, CleaningMaterial  # تأكد أن هذه النماذج موجودة

@bp.route("/warehouse_dashboard")
def warehouse_dashboard():
    assets_warehouse = Assetwarehouse.query.all()
    consumables = Consumable.query.all()
    cleaning = CleaningMaterial.query.all()
    return render_template("warehouse/warehouse_dashboard.html",
                           assets_warehouse=assets_warehouse,
                           consumables=consumables,
                           cleaning=cleaning)



@bp.route("/water_dashboard")
def water_dashboard():
    selected_housing = request.args.get("housing_units_id", "")
    selected_month = request.args.get("month", default=datetime.now().strftime("%Y-%m"))
    aggregation_type = request.args.get("aggregation_type", "monthly")

    # جلب البيانات باستخدام نفس المنطق الموجود في التقارير
    report_data, totals = calculate_consumption_data(selected_month, selected_housing)

    # إضافة بيانات المقيمين لكل وحدة كما في التقارير
    total_residents_today = 0
    total_residents_monthly = 0
    total_housing_costs = 0

    for row in report_data:
        row["avg_residents"] = get_total_residents(row["unit_id"], selected_month)
        row["total_residents"] = get_total_residents(row["unit_id"], selected_month, total=True)
        total_residents_today += row["avg_residents"]
        total_residents_monthly += row["total_residents"]
        total_housing_costs += row["total_cost"]
        row["cost_per_resident"] = row["total_cost"] / row["avg_residents"] if row["avg_residents"] else 0

        # تأكد من وجود dict للبنود حتى لا يحصل خطأ JSON
        if "per_item_costs" not in row or row["per_item_costs"] is None:
            row["per_item_costs"] = {}

    # حساب متوسط التكلفة للساكن
    avg_cost_per_resident = total_housing_costs / total_residents_today if total_residents_today else 0

    # ===========================
    # إعداد بيانات الرسم البياني للبنود
    # ===========================
    # جمع كل البنود الموجودة عبر كل السكنات
    all_expense_items = set()
    for row in report_data:
        all_expense_items.update((row.get("per_item_costs") or {}).keys())

    expense_chart_labels = sorted(all_expense_items)

    expense_chart_values = []
    for item in expense_chart_labels:
        total_item_cost = sum((row.get("per_item_costs") or {}).get(item, 0) for row in report_data)
        expense_chart_values.append(total_item_cost)

    # ===========================
    # إعداد بيانات الرسم البياني لتكاليف كل سكن
    # ===========================
    housing_names = [r["unit_name"] for r in report_data]
    housing_costs = [r["total_cost"] for r in report_data]

    # جلب كل السكنات للاختيار
    housings = HousingUnit.query.order_by(HousingUnit.name).all()

    return render_template(
        "water/water_dashboard.html",
        units=housings,
        selected_housing=selected_housing,
        selected_month=selected_month,
        aggregation_type=aggregation_type,
        housing_names=housing_names,
        housing_costs=housing_costs,
        total_residents_today=total_residents_today,
        total_residents_monthly=total_residents_monthly,
        total_housing_costs=total_housing_costs,
        avg_cost_per_resident=avg_cost_per_resident,
        report_data=report_data,  # ✅ إرسال البيانات كقائمة Python
        expense_chart_labels=expense_chart_labels,
        expense_chart_values=expense_chart_values,
        now=datetime.now()  # ✅ إضافة المتغير now
    )


# فلتر لإضافة فواصل الآلاف
# فلتر لإضافة فواصل الآلاف
@bp.app_template_filter()
def comma_format(value):
    try:
        return "{:,}".format(int(value))
    except (ValueError, TypeError):
        return value

# فلتر للعملة (ريال يمني)
@bp.app_template_filter()
def currency_format(value):
    try:
        return "{:,.0f} ريال".format(float(value))
    except (ValueError, TypeError):
        return value



@bp.route('/daily-resident', methods=['GET'])
def daily_resident_view():
    from datetime import datetime, timedelta, date
    from sqlalchemy import func, cast, Date, or_, and_, text

    month_str = request.args.get('month')
    if month_str:
        try:
            selected_month = datetime.strptime(month_str, '%Y-%m')
        except ValueError:
            selected_month = datetime.today().replace(day=1)
    else:
        selected_month = datetime.today().replace(day=1)

    first_day = selected_month.replace(day=1)
    if selected_month.month == 12:
        last_day = first_day.replace(year=selected_month.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        last_day = first_day.replace(month=selected_month.month + 1, day=1) - timedelta(days=1)

    days_in_month = []
    current_day = first_day.date()
    last_day_date = last_day.date()

    while current_day <= last_day_date:
        days_in_month.append(current_day)
        current_day += timedelta(days=1)

    print(f"🔍 [FINAL_SOLUTION] فحص شهر: {selected_month.strftime('%Y-%m')}")

    # ✅ تحديث الجلسة
    db.session.expire_all()

    # 🔧 الحل النهائي: استخدام استعلام SQL مباشر لدمج date و check_in
    sql_query = text("""
        SELECT 
            COALESCE(TRIM(hu.name), 'غير محدد') as housing_name,
            DATE(fa.date || ' ' || CASE 
                WHEN fa.check_in != '' AND fa.check_in != '-' THEN fa.check_in 
                ELSE '00:00:00' 
            END) as date_only,
            COUNT(fa.id) as count
        FROM fingerprint_archive fa
        JOIN employees e ON fa.employee_code = e.employee_code
        LEFT JOIN bed_assignments ba ON e.id = ba.employee_id 
            AND ba.active = 1 
            AND ba.start_date <= :last_day 
            AND (ba.end_date IS NULL OR ba.end_date >= :first_day)
        LEFT JOIN beds b ON ba.bed_id = b.id
        LEFT JOIN rooms r ON b.room_id = r.id
        LEFT JOIN housing_units hu ON r.housing_unit_id = hu.id
        WHERE DATE(fa.date) >= :first_day 
            AND DATE(fa.date) <= :last_day
        GROUP BY hu.name, date_only
        ORDER BY housing_name, date_only
    """)

    all_entries = db.session.execute(sql_query, {
        'first_day': first_day.date(),
        'last_day': last_day.date()
    }).fetchall()

    print(f"📊 [FINAL_SOLUTION] عدد السجلات: {len(all_entries)}")

    # 🔧 فحص خاص: عرض عينة من البيانات
    sample_entries = all_entries[:5]
    print(f"🔍 [FINAL_SOLUTION] عينة من البيانات:")
    for entry in sample_entries:
        print(f"   🏠 {entry.housing_name} - {entry.date_only}: {entry.count}")

    data_dict = {}
    for entry in all_entries:
        housing_name = entry.housing_name
        entry_date = entry.date_only

        # التأكد من أن entry_date هو date
        if isinstance(entry_date, datetime):
            entry_date = entry_date.date()
        elif isinstance(entry_date, str):
            try:
                entry_date = datetime.strptime(entry_date, '%Y-%m-%d').date()
            except ValueError:
                continue

        key = (housing_name, entry_date)
        data_dict[key] = entry.count

    housings_list = [
        'سكن المتعهدين',
        'سكن الفنيين',
        'سكن.ش. راس عيس',
        'سكن الهندسية',
        'سكن الادارين (C)',
        'سكن المشرفين (A)',
        'سكن الزائرين',
        'سكن ر . الاقسام (B)',
        'غير محدد'
    ]

    table_data = []

    for housing_name in housings_list:
        row = {"housing_unit": housing_name, "daily_counts": [], "total": 0}

        for day in days_in_month:
            count = data_dict.get((housing_name, day), 0)
            row["daily_counts"].append(count)
            row["total"] += count

        non_zero_days = sum(1 for count in row["daily_counts"] if count > 0)
        row["average"] = round(row["total"] / non_zero_days, 1) if non_zero_days else 0

        table_data.append(row)

    # حساب الإجماليات
    column_totals = [0] * len(days_in_month)
    for row in table_data:
        for i, count in enumerate(row["daily_counts"]):
            column_totals[i] += count

    grand_total = sum(row["total"] for row in table_data)
    grand_average = sum(row["average"] for row in table_data) / len(table_data) if table_data else 0
    grand_average = round(grand_average, 1)

    print(f"📊 [FINAL_SOLUTION] الإجمالي: {grand_total}, إجمالي أول يوم: {column_totals[0]}")

    return render_template(
        "water/daily_residents_form.html",
        table_data=table_data,
        days_in_month=days_in_month,
        selected_month=selected_month,
        column_totals=column_totals,
        grand_total=grand_total,
        grand_average=grand_average
    )


@bp.route("/expenses")
def expenseItem_index():
    items = ExpenseItem.query.order_by(ExpenseItem.name).all()
    housings = HousingUnit.query.order_by(HousingUnit.name).all()
    return render_template("water/expenses.html", items=items, housings=housings)

# إضافة بند جديد
@bp.route("/expense_items/add", methods=["POST"], endpoint="expenseItem_add")
def expenseItem_add():
    name = request.form.get("name")
    description = request.form.get("description")
    calculation_type = request.form.get("calculation_type")
    selected_housings = request.form.getlist("allowed_housings")

    if not name:
        flash("يرجى إدخال اسم البند", "warning")
        return redirect(url_for("main.expenseItem_index"))

    item = ExpenseItem(
        name=name,
        description=description,
        calculation_type=calculation_type
    )

    if selected_housings:
        item.housings = HousingUnit.query.filter(HousingUnit.id.in_(selected_housings)).all()

    db.session.add(item)
    db.session.commit()
    flash("تم إضافة البند بنجاح", "success")
    return redirect(url_for("main.expenseItem_index"))

# تعديل بند موجود
@bp.route("/expenseItem_edit/edit/<int:item_id>", methods=["POST"], endpoint="expenseItem_edit")
def expenseItem_edit(item_id):
    item = ExpenseItem.query.get_or_404(item_id)
    item.name = request.form.get("name")
    item.description = request.form.get("description")
    item.unit_price = request.form.get("unit_price", type=float)
    item.unit_qtr = request.form.get("unit_qtr")
    item.calculation_type = request.form.get("calculation_type")

    selected_housings = request.form.getlist("allowed_housings")
    if selected_housings:
        item.housings = HousingUnit.query.filter(HousingUnit.id.in_(selected_housings)).all()
    else:
        item.housings = []

    db.session.commit()
    flash("تم تحديث البند بنجاح", "success")
    return redirect(url_for("main.expenseItem_index"))

# حذف بند
@bp.route("/expense_items/delete/<int:item_id>", methods=["POST"], endpoint="expenseItem_delete")
def expenseItem_delete(item_id):
    item = ExpenseItem.query.get_or_404(item_id)

    # منع الحذف إذا مرتبط بالاستهلاك الشهري
    if item.consumptions and len(item.consumptions) > 0:
        flash("لا يمكن حذف هذا البند لأنه مرتبط بالاستهلاك الشهري", "warning")
        return redirect(url_for("main.expense_items"))

    db.session.delete(item)
    db.session.commit()
    flash("تم حذف البند بنجاح", "success")
    return redirect(url_for("main.expense_items"))


from datetime import datetime, timedelta
from sqlalchemy import func
def link_expense_items_to_housings():
    all_items = ExpenseItem.query.all()
    all_housings = HousingUnit.query.all()

    for item in all_items:
        # ربط كل البنود بالسكنات
        item.housings = all_housings

    db.session.commit()
    print("✅ تم ربط البنود بالسكنات بنجاح")

from datetime import datetime, timedelta
from sqlalchemy import func

def get_total_residents(housing_identifier, month_input, allowed_housings_item=None, total=False):

    """
    إرجاع متوسط أو إجمالي عدد الساكنين الشهري لسكن محدد
    total=False => ترجع المتوسط
    total=True  => ترجع الإجمالي
    """
    if isinstance(month_input, str):
        try:
            month = datetime.strptime(month_input, "%Y-%m")
        except ValueError:
            month = datetime.today().replace(day=1)
    elif isinstance(month_input, datetime):
        month = month_input
    else:
        month = datetime.today().replace(day=1)

    first_day = month.replace(day=1)
    last_day = (first_day.replace(month=month.month+1, day=1) - timedelta(days=1)) \
        if month.month != 12 else (first_day.replace(year=month.year+1, month=1, day=1) - timedelta(days=1))

    db.session.expire_all()

    query = db.session.query(func.count(FingerprintArchive.id)) \
        .join(Employee, FingerprintArchive.employee_code == Employee.employee_code) \
        .outerjoin(BedAssignment, Employee.id == BedAssignment.employee_id) \
        .outerjoin(Bed, BedAssignment.bed_id == Bed.id) \
        .outerjoin(Room, Bed.room_id == Room.id) \
        .outerjoin(HousingUnit, Room.housing_unit_id == HousingUnit.id) \
        .filter(FingerprintArchive.date >= first_day, FingerprintArchive.date <= last_day)

    if housing_identifier is not None:
        if isinstance(housing_identifier, int):
            query = query.filter(HousingUnit.id == housing_identifier)
        else:
            query = query.filter(func.trim(HousingUnit.name) == housing_identifier)

    daily_counts = query.group_by(FingerprintArchive.date).all()
    counts = [c[0] for c in daily_counts]

    if total:
        return sum(counts)  # إجمالي الساكنين
    else:
        non_zero_days = sum(1 for c in counts if c > 0)
        return round(sum(counts) / non_zero_days, 1) if non_zero_days else 0  # المتوسط

    form = HousingUnitForm()

    housings_list = HousingUnit.query.all()

    items_list = ExpenseItem.query.order_by(ExpenseItem.name).all()

    if form.validate_on_submit():
        if HousingUnit.query.filter_by(name=form.name.data).first():
            flash('اسم المستخدم موجود مسبقاً', 'warning')
            return render_template('water/monthly_consumption.html', form=form, users=users_list, items=items_list)
        # مثال: إضافة مستخدم
        housing_units = HousingUnit(
            name=form.name.data,
            username=form.username.data
        )
        # ربط المستخدم بالمناطق المختارة
        housing_units.items = HousingUnit.query.filter(HousingUnit.id.in_(form.item_ids.data)).all()
        db.session.add(user)
        db.session.commit()


from flask import request, render_template, flash, redirect, url_for
from sqlalchemy.orm import joinedload
from models import db, HousingUnit, ExpenseItem, MonthlyConsumption
@bp.route("/monthly_consumption", methods=["GET", "POST"])
def monthly_consumption():
    housings = HousingUnit.query.order_by(HousingUnit.name).all()
    housing_unit_id = request.form.get("housing_unit_id") if request.method == "POST" else None
    month = request.form.get("month") if request.method == "POST" else None

    data = []

    if housing_unit_id and month:
        housing_unit = HousingUnit.query.options(joinedload(HousingUnit.items)).get(int(housing_unit_id))

        consumptions = {
            mc.expense_item_id: mc
            for mc in MonthlyConsumption.query.filter_by(
                housing_unit_id=housing_unit_id,
                month=month
            ).all()
        }

        # تجهيز البيانات للعرض
        for item in housing_unit.items:
            mc = consumptions.get(item.id)
            unit_price = mc.unit_price if mc and mc.unit_price is not None else (item.unit_price or 0)
            qty = mc.qty if mc else 0

            if item.calculation_type == "normal":
                total_price = qty * unit_price
            elif item.calculation_type == "fixed":
                total_residents = get_total_residents(housing_unit.id, month, total=True)
                qty = total_residents  # الكمية دائماً = إجمالي الساكنين
                unit_price = mc.unit_price if mc and mc.unit_price is not None else (item.unit_price or 0)
                total_price = qty * unit_price

            elif item.calculation_type == "percentage":
                allowed_housings = item.housings
                total_residents = get_total_residents(housing_unit.id, month)
                overall_residents = sum(get_total_residents(h.id, month) for h in allowed_housings)
                ratio = total_residents / overall_residents if overall_residents else 0
                qty = ratio  # للعرض فقط
                total_price = (unit_price or 0) * ratio
            else:
                total_price = qty * unit_price

            data.append({
                "expense_item_id": item.id,
                "name": item.name,
                "unit_price": unit_price,
                "qty": qty,
                "calculation_type": item.calculation_type,
                "total_price": round(total_price, 2)
            })

        # حفظ الكميات والأسعار فقط للبنود القابلة للتعديل
        if request.method == "POST" and "save" in request.form:
            for item in housing_unit.items:
                if item.calculation_type == "percentage":
                    # لا نحفظ النسب لأنها محسوبة تلقائيًا
                    continue

                unit_price = request.form.get(f"unit_price_{item.id}")
                qty = request.form.get(f"qty_{item.id}")

                mc = MonthlyConsumption.query.filter_by(
                    housing_unit_id=housing_unit_id,
                    expense_item_id=item.id,
                    month=month
                ).first()

                unit_price = float(unit_price or 0)
                qty = float(qty or 0)

                if mc:
                    mc.unit_price = unit_price
                    mc.qty = qty
                else:
                    mc = MonthlyConsumption(
                        housing_unit_id=housing_unit_id,
                        expense_item_id=item.id,
                        month=month,
                        unit_price=unit_price,
                        qty=qty
                    )
                    db.session.add(mc)

            db.session.commit()
            flash("تم الحفظ بنجاح ✅", "success")
            return redirect(url_for("main.monthly_consumption", housing_unit_id=housing_unit_id, month=month))

    return render_template(
        "water/monthly_consumption.html",
        housings=housings,
        data=data,
        housing_unit_id=housing_unit_id,
        month=month
    )


from models import db, HousingUnit, ExpenseItem, MonthlyConsumption

# ===========================
# فلتر لإضافة فواصل الآلاف
@bp.app_template_filter()
def comma_format(value):
    try:
        return "{:,}".format(int(value))
    except (ValueError, TypeError):
        return value

# فلتر للعملة (ريال يمني)
@bp.app_template_filter()
def currency_format(value):
    try:
        return "{:,.0f} ريال".format(float(value))
    except (ValueError, TypeError):
        return value
# ===========================


# ===========================
# دالة لحساب بيانات الاستهلاك
# ===========================
# دالة لحساب بيانات الاستهلاك
def calculate_consumption_data(selected_month='', selected_housing=''):
    housings = HousingUnit.query.order_by(HousingUnit.name).all()
    items = ExpenseItem.query.order_by(ExpenseItem.name).all()

    # جلب الاستهلاك الشهري مسبقاً لكل السجلات
    records = MonthlyConsumption.query.filter_by(month=selected_month).all()
    # تسهيل البحث
    records_dict = {(r.housing_unit_id, r.expense_item_id): r for r in records}

    data = []
    totals_per_item = {item.name: 0 for item in items}
    overall_total = 0
    total_residents_sum = 0
    total_avg_residents_sum = 0

    for unit in housings:
        if selected_housing and str(unit.id) != str(selected_housing):
            continue

        # حساب بيانات المقيمين لهذه الوحدة
        avg_residents = get_total_residents(unit.id, selected_month)
        total_residents = get_total_residents(unit.id, selected_month, total=True)

        row = {
            "unit_id": unit.id,
            "unit_name": unit.name,
            "total_cost": 0,
            "per_item_costs": {},
            "avg_residents": avg_residents,
            "total_residents": total_residents
        }

        for item in items:
            # جلب السجل المحفوظ إن وجد
            rec = records_dict.get((unit.id, item.id))

            if rec:
                unit_price = rec.unit_price if rec.unit_price is not None else item.unit_price or 0
                qty = rec.qty
            else:
                unit_price = item.unit_price or 0
                qty = 0

            value = 0

            if item.calculation_type == "normal":
                value = qty * unit_price
            elif item.calculation_type == "fixed":
                total_residents = get_total_residents(unit.id, selected_month, total=True)
                qty = total_residents  # الكمية تعرض دائماً إجمالي الساكنين
                unit_price = rec.unit_price if rec and rec.unit_price is not None else item.unit_price or 0
                value = qty * unit_price
            elif item.calculation_type == "percentage":
                allowed_housings = item.housings
                total_residents = get_total_residents(unit.id, selected_month)
                overall_residents = sum(get_total_residents(h.id, selected_month) for h in allowed_housings)
                ratio = total_residents / overall_residents if overall_residents else 0
                if not rec:  # إذا لم يتم حفظه مسبقًا
                    qty = ratio
                    value = unit_price * ratio
                else:  # استخدم القيم المحفوظة
                    value = qty * unit_price
            else:
                value = qty * unit_price

            row[item.name] = value
            row["per_item_costs"][item.name] = value
            row["total_cost"] += value
            totals_per_item[item.name] += value

        # حساب تكلفة الساكن الواحد
        row["cost_per_resident"] = row["total_cost"] / avg_residents if avg_residents else 0

        overall_total += row["total_cost"]
        total_residents_sum += total_residents
        total_avg_residents_sum += avg_residents

        data.append(row)

    # حساب الإجماليات الكلية
    total_cost_per_resident = overall_total / total_avg_residents_sum if total_avg_residents_sum else 0

    totals = {
        "overall_cost": overall_total,
        "per_item": totals_per_item,
        "total_residents": total_residents_sum,
        "total_avg_residents": total_avg_residents_sum,
        "total_cost_per_resident": total_cost_per_resident
    }

    return data, totals


# ===========================
# واجهة استهلاك
@bp.route('/consumption', methods=['GET', 'POST'])
def consumption():
    selected_month = request.args.get("month", default=datetime.now().strftime("%Y-%m"))
    selected_housing = request.args.get("housing_units_id", default="")

    data, totals = calculate_consumption_data(selected_month, selected_housing)

    housing_units = HousingUnit.query.order_by(HousingUnit.name).all()
    expense_items = ExpenseItem.query.order_by(ExpenseItem.name).all()

    return render_template(
        "water/consumption.html",
        data=data,
        items=expense_items,
        totals=totals,
        selected_month=selected_month,
        selected_housing=selected_housing,
        housing_units=housing_units
    )


# ===========================
# تقرير الاستهلاك
@bp.route('/consumption/report')
def consumption_report():
    selected_housing = request.args.get("housing_units_id", type=int)
    selected_month = request.args.get("month", default=datetime.now().strftime("%Y-%m"))

    data, totals = calculate_consumption_data(selected_month, selected_housing)

    report_data = []
    for row in data:
        report_data.append({
            "unit_name": row["unit_name"],
            "month": selected_month,
            "total_cost": row["total_cost"],
            "total_residents": row["total_residents"],
            "avg_residents": row["avg_residents"],
            "cost_per_resident": row["cost_per_resident"]
        })

    return render_template(
        "water/consumption_report.html",
        data=report_data,
        totals=totals,
        units=HousingUnit.query.order_by(HousingUnit.name).all(),
        selected_housing=str(selected_housing) if selected_housing else "",
        selected_month=selected_month
    )
from sqlalchemy import extract, func
from models import db, DailyResident, MonthlyResidentAverage
from flask import render_template
from calendar import month_name
from sqlalchemy import text
from sqlalchemy import text
from sqlalchemy import func

from sqlalchemy import func
from flask import render_template
from sqlalchemy import func
from datetime import datetime
from models import db, DailyResident, HousingUnit

@bp.route('/average_residents')
def average_residents():
    """
    عرض متوسط عدد الساكنين لكل وحدة سكن في كل شهر
    مباشرة من جدول daily_resident
    """
    # جلب المتوسطات الشهرية مباشرة من الجدول اليومي
    results = db.session.query(
        HousingUnit.name.label('housing_name'),
        func.strftime('%Y-%m', DailyResident.date).label('year_month'),
        func.avg(DailyResident.residents).label('average_residents')
    ).join(HousingUnit, HousingUnit.id == DailyResident.housing_unit_id) \
     .group_by(HousingUnit.name, 'year_month') \
     .order_by(HousingUnit.name, 'year_month') \
     .all()

    # تحويل النتائج لقائمة dict لسهولة العرض في القالب
    data = [{
        'housing_name': row.housing_name,
        'year_month': row.year_month,
        'average_residents': round(row.average_residents, 1) if row.average_residents else 0
    } for row in results]

    return render_template('water/average_residents.html', data=data)

from sqlalchemy import func
from sqlalchemy import func

def update_monthly_residents_avg():
    # حساب المتوسط لكل سكن ولكل شهر
    averages = (
        db.session.query(
            DailyResident.housing_unit_id,
            func.strftime('%Y-%m', DailyResident.date).label('year_month'),
            func.avg(DailyResident.residents).label('avg_residents')
        )
        .group_by(DailyResident.housing_unit_id, 'year_month')
        .all()
    )

    for housing_unit_id, year_month, avg_residents in averages:
        # تحقق إذا كان السجل موجود مسبقاً
        record = MonthlyResidentAverage.query.filter_by(
            housing_unit_id=housing_unit_id,
            year_month=year_month
        ).first()

        if record:
            # تحديث المتوسط
            record.average_residents = round(avg_residents, 2)
        else:
            # إضافة سجل جديد
            new_record = MonthlyResidentAverage(
                housing_unit_id=housing_unit_id,
                year_month=year_month,
                average_residents=round(avg_residents, 2)
            )
            db.session.add(new_record)

    db.session.commit()
@bp.route('/update_monthly_avg')
def update_monthly_avg():
    update_monthly_residents_avg()
    return "تم تحديث المتوسطات الشهرية بنجاح"


from sqlalchemy import func
from sqlalchemy import extract
from models import MonthlyConsumption, HousingUnit
from sqlalchemy.sql import func
from datetime import datetime
from sqlalchemy.sql import func
from datetime import datetime

from flask import make_response
from fpdf import FPDF
from flask import make_response, send_file
from fpdf import FPDF
from io import BytesIO

@bp.route('/export_consumption_pdf')
def export_consumption_pdf():
    consumptions = MonthlyConsumption.query.all()
    selected_month = request.args.get('month')
    selected_housing = request.args.get('housing_unit_id')

    if selected_month:
        consumptions = [c for c in consumptions if c.month == selected_month]
    if selected_housing:
        consumptions = [c for c in consumptions if str(c.housing_unit_id) == selected_housing]

    pdf = FPDF()
    pdf.add_page()
    pdf.add_font('Amiri', '', r'D:\ghith\NEW\costing\fonts\Amiri-Regular.ttf', uni=True)
    pdf.set_font('Amiri', '', 14)
    pdf.cell(0, 10, "تقرير استهلاك الماء والكهرباء", 0, 1, 'C')

    for c in consumptions:
        total = 0
        if c.water_qty and c.water_price:
            total += c.water_qty * c.water_price
        if c.electricity_qty and c.electricity_price:
            total += c.electricity_qty * c.electricity_price
        if c.other_qty and c.other_price:
            total += c.other_qty * c.other_price

        # تأكد من وجود اسم السكن قبل الاستخدام لتجنب الخطأ
        housing_name = c.housing_unit.name if c.housing_unit else "غير معروف"
        line = f"{housing_name} - {c.month} - {total:.2f} ريال"
        pdf.cell(0, 10, line, 0, 1)

    pdf_bytes = bytes(pdf.output(dest='S'))  # تحويل bytearray إلى bytes
    response = make_response(pdf_bytes)
    response.headers.set('Content-Disposition', 'attachment', filename='report.pdf')
    response.headers.set('Content-Type', 'application/pdf')
    return response


@bp.route('/export_consumption_excel')
def export_consumption_excel():
    consumptions = MonthlyConsumption.query.all()
    selected_month = request.args.get('month')
    selected_housing = request.args.get('housing_unit_id')

    if selected_month:
        consumptions = [c for c in consumptions if c.month == selected_month]
    if selected_housing:
        consumptions = [c for c in consumptions if str(c.housing_unit_id) == selected_housing]

    data = []
    for c in consumptions:
        total = 0
        if c.water_qty and c.water_price:
            total += c.water_qty * c.water_price
        if c.electricity_qty and c.electricity_price:
            total += c.electricity_qty * c.electricity_price
        if c.other_qty and c.other_price:
            total += c.other_qty * c.other_price

        housing_name = c.housing_unit.name if c.housing_unit else "غير معروف"
        data.append({
            'السكن': housing_name,
            'الشهر': c.month,
            'كمية الماء': c.water_qty or 0,
            'سعر الماء': c.water_price or 0,
            'كمية الكهرباء': c.electricity_qty or 0,
            'سعر الكهرباء': c.electricity_price or 0,
            'الوصف الآخر': c.other_desc or '',
            'كمية أخرى': c.other_qty or 0,
            'سعر أخرى': c.other_price or 0,
            'الإجمالي': total
        })

    df = pd.DataFrame(data)

    # تحسين تنسيق الأعمدة أو التعامل مع الفارغة إذا أردت

    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='تقرير')
        workbook = writer.book
        worksheet = writer.sheets['تقرير']

        # يمكنك هنا تحسين التنسيق مثل عرض الأعمدة تلقائيًا
        for idx, col in enumerate(df.columns):
            series = df[col]
            max_len = max((
                series.astype(str).map(len).max(),
                len(str(series.name))
            )) + 2  # مسافة إضافية
            worksheet.set_column(idx, idx, max_len)

    output.seek(0)

    return send_file(output,
                     download_name="report.xlsx",
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@bp.route('/report')
def report():
    return render_template('water/report.html')


# فلتر لإضافة فواصل الآلاف
@bp.app_template_filter()
def comma_format(value):
    try:
        return "{:,}".format(int(value))
    except (ValueError, TypeError):
        return value


# فلتر للعملة (ريال يمني)
@bp.app_template_filter()
def currency_format(value):
    try:
        return "{:,.0f} ريال".format(float(value))
    except (ValueError, TypeError):
        return value


# فلتر جديد لنسبة الحضور
@bp.app_template_filter()
def percentage_format(value):
    try:
        return "{:.1f}%".format(float(value))
    except (ValueError, TypeError):
        return value


from flask import Blueprint, jsonify
from datetime import datetime, timedelta, date
from flask import request, render_template
from datetime import datetime, timedelta
from models import HousingUnit, FingerprintArchive, Employee, Company, BedAssignment, Bed, Room

@bp.route('/employee_attendance_report', methods=['GET'])
def employee_attendance_report():
    from datetime import datetime, timedelta
    from sqlalchemy import func

    # تحديد الشهر المختار
    month_str = request.args.get('month')
    if month_str:
        try:
            selected_month = datetime.strptime(month_str, '%Y-%m')
        except ValueError:
            selected_month = datetime.today().replace(day=1)
    else:
        selected_month = datetime.today().replace(day=1)

    # كل أيام الشهر
    first_day = selected_month.replace(day=1)
    if selected_month.month == 12:
        last_day = first_day.replace(year=selected_month.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        last_day = first_day.replace(month=selected_month.month + 1, day=1) - timedelta(days=1)
    days_in_month = [first_day + timedelta(days=i) for i in range((last_day - first_day).days + 1)]

    # فلترة نوع الخدمة
    selected_service_type = request.args.get('service_type', None)
    service_types_query = db.session.query(Employee.service_type).distinct().all()
    service_types = [stype[0] for stype in service_types_query if stype[0]]

    # إنشاء dict لجميع الشركات أولاً لضمان ظهورها
    all_companies = db.session.query(Employee.company_id, Employee.company_name).distinct().all()
    data_dict = {}
    for comp in all_companies:
        comp_id = comp.company_id if comp.company_id is not None else f"none_{comp.company_name}"
        comp_name = comp.company_name or "غير محدد"
        data_dict[comp_id] = {"id": comp_id, "name": comp_name, "employees": {}}

    # إضافة جميع الموظفين النشطين لكل شركة
    all_employees_query = db.session.query(
        Employee.id.label("employee_id"),
        Employee.name.label("employee_name"),
        Employee.job_title,
        Employee.employee_code,
        Employee.company_id,
        Employee.company_name,
        Employee.service_type
    )
    if selected_service_type:
        all_employees_query = all_employees_query.filter(Employee.service_type == selected_service_type)
    all_employees = all_employees_query.all()

    for emp in all_employees:
        emp_id = emp.employee_id
        emp_name = emp.employee_name
        emp_job = emp.job_title or "-"
        comp_id = emp.company_id if emp.company_id is not None else f"none_{emp.company_name}"

        if comp_id not in data_dict:
            data_dict[comp_id] = {"id": comp_id, "name": emp.company_name or "غير محدد", "employees": {}}

        if emp_id not in data_dict[comp_id]["employees"]:
            data_dict[comp_id]["employees"][emp_id] = {
                "name": emp_name,
                "job_title": emp_job,
                "attendance": [0] * len(days_in_month),
                "present_days": 0,
                "absent_days": len(days_in_month),
                "attendance_rate": 0
            }

    # استعلام سجلات الحضور للشهر
    all_entries_query = db.session.query(
        Employee.id.label("employee_id"),
        Employee.company_id,
        Employee.company_name,
        func.date(FingerprintArchive.date).label("date_only")
    ).join(Employee, FingerprintArchive.employee_code == Employee.employee_code)

    # إذا تم تحديد نوع خدمة، نفلتر سجلات الحضور
    if selected_service_type:
        all_entries_query = all_entries_query.filter(Employee.service_type == selected_service_type)

    all_entries = all_entries_query.filter(
        func.substr(FingerprintArchive.date, 1, 7) == selected_month.strftime('%Y-%m')
    ).all()

    # تحديث الحضور
    for entry in all_entries:
        emp_id = entry.employee_id
        comp_id = entry.company_id if entry.company_id is not None else f"none_{entry.company_name}"

        # التأكد من وجود الموظف في dict قبل التحديث
        if comp_id not in data_dict:
            data_dict[comp_id] = {"id": comp_id, "name": entry.company_name or "غير محدد", "employees": {}}

        if emp_id not in data_dict[comp_id]["employees"]:
            # إضافة الموظف إذا لم يكن موجود
            data_dict[comp_id]["employees"][emp_id] = {
                "name": "-",
                "job_title": "-",
                "attendance": [0] * len(days_in_month),
                "present_days": 0,
                "absent_days": len(days_in_month),
                "attendance_rate": 0
            }

        # تحويل التاريخ
        date_only = None
        if entry.date_only:
            if isinstance(entry.date_only, str):
                try:
                    date_only = datetime.strptime(entry.date_only[:10], "%Y-%m-%d")
                except ValueError:
                    date_only = None
            else:
                date_only = entry.date_only

        if date_only:
            day_index = (date_only - first_day).days
            if 0 <= day_index < len(days_in_month):
                emp_data = data_dict[comp_id]["employees"][emp_id]
                if emp_data["attendance"][day_index] == 0:
                    emp_data["attendance"][day_index] = 1
                    emp_data["present_days"] += 1
                    emp_data["absent_days"] -= 1

    # تجهيز table_data
    table_data = []
    grand_daily_totals = [0] * len(days_in_month)
    total_employees = 0
    grand_total_present = 0
    grand_total_absent = 0

    for comp in data_dict.values():
        company_emps = list(comp["employees"].values())
        for emp in company_emps:
            emp["present_days"] = sum(emp["attendance"])
            emp["absent_days"] = len(days_in_month) - emp["present_days"]
            emp["attendance_rate"] = round((emp["present_days"] / len(days_in_month)) * 100, 1) if len(days_in_month) > 0 else 0

        daily_totals = [sum(emp["attendance"][i] for emp in company_emps) for i in range(len(days_in_month))]
        total_present = sum(emp["present_days"] for emp in company_emps)
        total_absent = sum(emp["absent_days"] for emp in company_emps)
        attendance_rate = round((total_present / (len(company_emps) * len(days_in_month))) * 100, 1) if company_emps else 0

        table_data.append({
            "id": comp["id"],
            "name": comp["name"],
            "employees": company_emps,
            "daily_totals": daily_totals,
            "total_present": total_present,
            "total_absent": total_absent,
            "attendance_rate": attendance_rate
        })

        for i in range(len(days_in_month)):
            grand_daily_totals[i] += daily_totals[i]
        total_employees += len(company_emps)
        grand_total_present += total_present
        grand_total_absent += total_absent

    grand_attendance_rate = round((grand_total_present / (total_employees * len(days_in_month))) * 100, 1) if total_employees > 0 else 0

    return render_template(
        "water/daily_employees.html",
        table_data=table_data,
        days_in_month=days_in_month,
        selected_month=selected_month,
        selected_service_type=selected_service_type,
        service_types=service_types,
        grand_daily_totals=grand_daily_totals,
        total_employees=total_employees,
        grand_total_present=grand_total_present,
        grand_total_absent=grand_total_absent,
        grand_attendance_rate=grand_attendance_rate,
        employees_count=total_employees,
        companies_count=len(table_data),
        grand_total_attendance=grand_total_present,
        now=datetime.now()
    )

from models import work_type_schedule

@bp.route('/employee_attendance')
def employee_attendance():
    # الحصول على معاملات الفلترة
    # جدول أنوع الدوام: (أيام دوام, أيام إجازة)
    work_type_schedule = {
        "YCSR-A": (4, 3),
        "YCSR-B": (4, 3),
        "YCSR-C": (4, 3),
        "YCSR-D": (4, 3),
        "YCSR-c": (4, 3),
        "YCSRW1": (7, 7),
        "YCSRW2": (7, 7),
        "YCSRK2": (14, 7),
        "YCSRK1": (14, 7),
        "YCSRSH1": (6, 1),
        "N_YCSRE2": (6, 1),# التاكد من نوع الدوام
        "ycsrsc": (6, 1),  # التاكد من نوع الدوام
        "YCSRE1": (6, 1),  # التاكد من نوع الدوام
    "YCSR-E": (6, 1),  # إضافتك الجديدة: 6 أيام دوام، 1 إجازة

        "N_PR28H1": (28, 28),
        "N_PR28H2": (28, 28),
        "N_PR28H3": (28, 28),
        "N_PR28H4": (28, 28),
        "SHIFT2": (42, 14),
        "YCSRK3": (42, 14),
        "NORMYS4": (5, 2),
        "NORMALY2": (6, 1),
        "NORMALY3": (7, 0),
        "NORMYS5": (4, 3),
        "NORMYS6": (4, 3),
        "NORMYS7": (42, 21),
        "غير محدد": (6, 1),
        "NORMYS12": (6, 1),# التاكد من نوع الدوام

        "ورديات": (6, 1),
        "NORMYS2": (6, 1)

    }

    attendance_date = request.args.get('attendance_date', datetime.now().strftime('%Y-%m-%d'))
    work_type_filter = request.args.get('work_type', '')
    housing_filter = request.args.get('housing_location', '')

    try:
        selected_date = datetime.strptime(attendance_date, '%Y-%m-%d').date()
    except:
        selected_date = datetime.now().date()

    # بناء الاستعلام الأساسي
    query = BedAssignment.query.join(Employee).join(Bed).join(Room).join(HousingUnit)

    # تطبيق الفلترة
    if work_type_filter:
        query = query.filter(Employee.work_type == work_type_filter)

    if housing_filter:
        query = query.filter(HousingUnit.name == housing_filter)

    # الفلترة حسب التاريخ
    query = query.filter(
        BedAssignment.start_date <= selected_date,
        (BedAssignment.end_date == None) | (BedAssignment.end_date >= selected_date),
        BedAssignment.active == True
    )

    assignments = query.all()

    # جمع الإحصائيات
    stats = {
        'total_employees': len(assignments),
        'active_employees': sum(1 for a in assignments if
                                a.status == 'نشط'),  # استدعاء الحالة من property الموجودة في BedAssignment
        'on_leave': sum(1 for a in assignments if
                        a.status == 'إجازة'),
        'occupancy_rate': round((len(assignments) / Bed.query.count()) * 100, 2) if Bed.query.count() > 0 else 0
    }

    return render_template(
        'water/work_type.html',
        assignments=assignments,
        selected_date=selected_date,
        work_types=list(work_type_schedule.keys()),
        housing_locations=[h[0] for h in db.session.query(HousingUnit.name).distinct().all()],
        work_type_schedule=work_type_schedule,
        stats=stats
    )


@bp.route('/smart_attendance_report', methods=['GET'])
def smart_attendance_report():
    from datetime import datetime, date
    from sqlalchemy import or_, text, bindparam

    work_type_schedule = {
        "YCSR-A": (4, 3),  # ورديات من السبت الى الثلاثاء
        "YCSR-B": (4, 3),  # اداري من الاحد الى الاربعاء
        "YCSR-C": (4, 3),  # ورديات من الاثنين الى الخميس
        "YCSR-D": (4, 3),  # ورديات الاثنين والثلاثاء اجازة
        "YCSR-c": (4, 3),  # ورديات من الاربعاء الى السبت
        "YCSRW1": (7, 7),  # اسبوع * اسبوع
        "YCSRW2": (7, 7),  # اسبوع * اسبوع
        "YCSRK2": (14, 7),  # اسبوعين * اسبوع
        "YCSRK1": (14, 7),  # اسبوعين * اسبوع
        "YCSRSH1": (6, 1),  # ورديات عادية
        "N_PR28H1": (28, 28),  # شهر *شهر
        "N_PR28H2": (28, 28),  # شهر *شهر
        "N_PR28H3": (28, 28),  # شهر *شهر
        "N_PR28H4": (28, 28),  # شهر *شهر
        "SHIFT2": (42, 14),
        "YCSRK3": (42, 14),  # 6اسابيع * 3اسابيع
        "NORMYS4": (5, 2),  # اداري من الاحد الى الخميس
        "NORMALY2": (6, 1),
        "NORMALY3": (7, 0),
        "NORMYS5": (4, 3),  # اداري من السبت الى الثلاثاء
        "NORMYS6": (4, 3),  # اداري من الاثنين الى الخميس
        "NORMYS7": (42, 21),  # 6اسابيع * 3اسابيع
        "N_YCSRE2": (6, 1),
        "غير محدد": (6, 1),
        "ورديات": (6, 1),
        "ycsrsc": (6, 1),
        "YCSRE1": (6, 1),
        "NORMYS2": (6, 1)  # ورديات المبيعات

    }

    # معاملات الفلترة
    attendance_date_str = request.args.get('attendance_date', datetime.now().strftime('%Y-%m-%d'))
    work_type_filter = request.args.get('work_type', '')
    housing_filter = request.args.get('housing_location', '')
    service_type_filter = request.args.get('service_type', '')
    report_type = request.args.get('report_type', 'daily')

    # تحويل السلسلة إلى تاريخ
    try:
        selected_date = datetime.strptime(attendance_date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = datetime.now().date()

    selected_month = datetime(selected_date.year, selected_date.month, 1)
    days_in_month = []
    # 🔹 جلب معلومات الموظفين مع السكن (قبل سحب البصمات)
    # 🔹 جلب أكواد الموظفين من الاستعلام الرئيسي (بعد تحديد الفلاتر)
    employees_query = db.session.query(
        Employee.id,
        Employee.name,
        Employee.employee_code,
        Employee.job_title,
        Employee.work_type,
        Employee.service_type,
        Employee.company_name,
        BedAssignment.start_date,
        BedAssignment.end_date,
        HousingUnit.name.label('housing_name'),
        Room.room_number,
        Bed.bed_number
    ).join(BedAssignment, Employee.id == BedAssignment.employee_id) \
        .join(Bed, BedAssignment.bed_id == Bed.id) \
        .join(Room, Bed.room_id == Room.id) \
        .join(HousingUnit, Room.housing_unit_id == HousingUnit.id) \
        .filter(BedAssignment.active == True) \
        .filter(BedAssignment.start_date <= selected_date) \
        .filter(or_(BedAssignment.end_date == None, BedAssignment.end_date >= selected_date))

    # تطبيق فلاتر إضافية
    if work_type_filter:
        employees_query = employees_query.filter(Employee.work_type == work_type_filter)
    if housing_filter:
        employees_query = employees_query.filter(HousingUnit.name == housing_filter)
    if service_type_filter:
        employees_query = employees_query.filter(Employee.service_type == service_type_filter)

    employees_data = employees_query.all()

    # 🔹 جمع أكواد الموظفين
    employee_codes = [str(emp.employee_code).strip() for emp in employees_data]

    # 🔹 جلب البصمات من جدول FingerprintArchive
    fingerprint_dict = {}
    if employee_codes:
        all_fingerprints = (
            db.session.query(
                FingerprintArchive.employee_code,
                FingerprintArchive.check_in,
                FingerprintArchive.check_out,
                FingerprintArchive.date
            )
            .filter(FingerprintArchive.employee_code.in_(employee_codes))
            .filter(FingerprintArchive.date == selected_date)
            .all()
        )

        for fp in all_fingerprints:
            emp_code = str(fp.employee_code).strip()
            if emp_code not in fingerprint_dict:
                fingerprint_dict[emp_code] = []
            fingerprint_dict[emp_code].append({
                "check_in": fp.check_in,
                "check_out": fp.check_out,
                "date": fp.date
            })


        for fp in all_fingerprints:
            emp_code = fp[0]
            if emp_code not in fingerprint_dict:
                fingerprint_dict[emp_code] = []
            fingerprint_dict[emp_code].append({
                "check_in": fp[1],
                "check_out": fp[2],
                "date": fp[3]
            })

    # 🔹 الآن يمكن استدعاء employees_data
    employees_data = employees_query.all()

    # 🔹 معالجة البيانات النهائية
    report_data = []
    daily_summary = {'total_employees': 0, 'present_by_fingerprint': 0, 'absent': 0, 'on_leave': 0, 'undefined': 0}

    for emp in employees_data:
        safe_work_type = emp.work_type or "غير محدد"
        work_days, off_days = work_type_schedule.get(safe_work_type, (6, 1))

        start_date_obj = emp.start_date if isinstance(emp.start_date, date) else None

        # حالة الدوام
        schedule_status = "غير محدد"
        if start_date_obj:
            if selected_date < start_date_obj:
                schedule_status = "لم يبدأ الدوام"
            else:
                days_diff = (selected_date - start_date_obj).days
                cycle_length = work_days + off_days
                day_in_cycle = days_diff % cycle_length
                schedule_status = "نشط" if day_in_cycle < work_days else "إجازة"

        # حالة البصمة
        has_fingerprint = emp.employee_code in fingerprint_dict and len(fingerprint_dict[emp.employee_code]) > 0
        fingerprint_status = "مبصم" if has_fingerprint else "غير مبصم"

        # الحالة النهائية
        if has_fingerprint:
            final_status = "حاضر"
            status_class = "present"
        elif schedule_status == "نشط":
            final_status = "غائب"
            status_class = "absent"
        elif schedule_status == "إجازة":
            final_status = "إجازة"
            status_class = "on-leave"
        else:
            final_status = schedule_status
            status_class = "undefined"

        daily_summary['total_employees'] += 1
        if final_status == "حاضر":
            daily_summary['present_by_fingerprint'] += 1
        elif final_status == "غائب":
            daily_summary['absent'] += 1
        elif final_status == "إجازة":
            daily_summary['on_leave'] += 1
        else:
            daily_summary['undefined'] += 1

        report_data.append({
            'id': emp.id,
            'name': emp.name or "غير محدد",
            'employee_code': emp.employee_code,
            'job_title': emp.job_title or "-",
            'work_type': safe_work_type,
            'service_type': emp.service_type or "-",
            'company_name': emp.company_name or "غير محدد",
            'housing_name': emp.housing_name or "لا يوجد",
            'room_number': emp.room_number or "-",
            'bed_number': emp.bed_number or "-",
            'start_date': start_date_obj,
            'schedule_status': schedule_status,
            'fingerprint_status': fingerprint_status,
            'final_status': final_status,
            'status_class': status_class,
            'fingerprints': fingerprint_dict.get(emp.employee_code, []),
            'work_days': work_days,
            'off_days': off_days,
            'has_fingerprint': has_fingerprint
        })

    # قوائم الفلاتر
    work_types = list(work_type_schedule.keys())
    housing_locations = [h[0] for h in db.session.query(HousingUnit.name).distinct().all() if h[0]]
    service_types = [st[0] for st in db.session.query(Employee.service_type).distinct().all() if st[0]]

    return render_template(
        'water/work_type_daily.html',
        report_data=report_data,
        daily_summary=daily_summary,
        selected_date=selected_date,
        selected_month=selected_month,
        days_in_month=days_in_month,
        work_types=work_types,
        housing_locations=housing_locations,
        service_types=service_types,
        work_type_schedule=work_type_schedule,
        report_type=report_type,
        now=datetime.now()
    )
